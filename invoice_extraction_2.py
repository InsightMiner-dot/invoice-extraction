import fitz  # PyMuPDF
import base64
import instructor
import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
from pydantic import BaseModel, Field
from typing import List, Optional

# ==============================================================
# 1. Define Data Schema 
# ==============================================================

class LineItem(BaseModel):
    page_number: Optional[int] = Field(None, description="The page number (starting at 1) where this item was found")
    material: Optional[str] = Field(None, description="The material code, part number, or SKU")
    description: str = Field(description="The name or description of the item")
    quantity: Optional[float] = Field(None, description="The number of items purchased")
    uom: Optional[str] = Field(None, description="Unit of Measure (e.g., EA, LBS, KG)")
    uom_confidence: Optional[str] = Field(None, description="Rate confidence in UOM extraction: 'High', 'Medium', or 'Low'")
    unit_price: Optional[float] = Field(None, description="The price of a single unit")
    line_total: float = Field(description="The total cost for this specific line item. Look for 'Amount', 'Total', or 'Extended Price'.")

class InvoiceData(BaseModel):
    vendor_name: str = Field(description="The name of the company issuing the invoice")
    invoice_number: Optional[str] = Field(None, description="The unique invoice ID or number")
    invoice_number_confidence: Optional[str] = Field(None, description="Rate confidence in Invoice Number extraction: 'High', 'Medium', or 'Low'")
    date: Optional[str] = Field(None, description="The date the invoice was issued")
    vendor_address: Optional[str] = Field(None, description="The full address of the vendor")
    bill_to: Optional[str] = Field(None, description="The full 'Bill To' or 'Sold To' address")
    ship_to: Optional[str] = Field(None, description="The 'Ship To' address.")
    remit_to: Optional[str] = Field(None, description="The 'Remit To' address")
    origin: Optional[str] = Field(None, description="The origin address ('From', 'Ship From', 'Pickup', 'Generator').")
    origin_confidence: Optional[str] = Field(None, description="Rate confidence in Origin extraction: 'High', 'Medium', or 'Low'")
    destination: Optional[str] = Field(None, description="The destination address ('To', 'Deliver To', 'Consignee').")
    destination_confidence: Optional[str] = Field(None, description="Rate confidence in Destination extraction: 'High', 'Medium', or 'Low'")
    currency: Optional[str] = Field(None, description="The 3-letter currency code (e.g., USD, CAD, EUR)")
    tax_amount: Optional[float] = Field(0.0, description="The total tax amount. Default to 0.0 if not explicitly found.")
    shipping_handling: Optional[float] = Field(0.0, description="Shipping/Freight charges. Default to 0.0 if not explicitly found.")
    total_amount: float = Field(description="The final total amount charged on the invoice")
    total_amount_confidence: Optional[str] = Field(None, description="Rate confidence in Total Amount extraction: 'High', 'Medium', or 'Low'")
    line_items: List[LineItem] = Field(description="A list of all individual items purchased across all pages")

# ==============================================================
# 2. Setup Azure OpenAI Client
# ==============================================================

azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
azure_api_key = os.getenv("AZURE_OPENAI_KEY")
azure_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")
azure_api_version = os.getenv("AZURE_OPENAI_API_VERSION")

client = instructor.from_openai(
    AzureOpenAI(azure_endpoint=azure_endpoint, api_key=azure_api_key, api_version=azure_api_version)
)

# ==============================================================
# 3. PDF to Image Conversion & Extraction
# ==============================================================

def pdf_to_base64_images(pdf_path: str) -> List[str]:
    doc = fitz.open(pdf_path)
    base64_images = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
    return base64_images

def extract_invoice_data(pdf_path: str) -> InvoiceData:
    base64_images = pdf_to_base64_images(pdf_path)
    content_array = [{"type": "text", "text": "Extract data from this multi-page invoice. If a field is missing, leave it null. Evaluate confidence ('High', 'Medium', 'Low'). Note the page number (1, 2, etc.) for each line item."}]
    
    for img_base64 in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}})

    return client.chat.completions.create(
        model=azure_deployment, 
        response_model=InvoiceData, 
        messages=[
            {"role": "system", "content": "You are an expert accountant. Extract data accurately and provide honest confidence scores."},
            {"role": "user", "content": content_array}
        ]
    )

# ==============================================================
# 4. Excel Setup Utility
# ==============================================================

def setup_excel_workbook():
    """Creates a new workbook with 2 sheets and formatted headers."""
    wb = openpyxl.Workbook()
    
    # Setup Sheet 1: Details
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Invoice Number", "Inv# Confidence",
        "Origin", "Origin Conf", "Destination", "Dest Conf", "Bill To", "Ship To", "Remit To",
        "Currency", "Material", "Description", "Quantity", "UOM", "UOM Conf", 
        "Unit Price", "Line Total", "Invoice Total", "Total Conf"
    ]
    ws_details.append(details_headers)
    
    # Setup Sheet 2: QC Summary
    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "File Name", "Invoice Number", "Status", "Extracted Total", 
        "Calculated Sum (Lines + Tax + Ship)", "Variance", "Currency"
    ]
    ws_qc.append(qc_headers)
    
    # Apply Styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2" # Freeze top row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    return wb, ws_details, ws_qc

# ==============================================================
# 5. Main Execution & Folder Processing
# ==============================================================

if __name__ == "__main__":
    # ---> DIRECTORY CONFIGURATION <---
    input_folder = r"C:\path\to\your\invoices"  # Folder containing all PDFs
    review_folder = os.path.join(input_folder, "Needs_Manual_Review")
    output_excel = os.path.join(input_folder, "Extracted_Invoices_Master.xlsx")
    
    # Create the review folder if it doesn't exist
    if not os.path.exists(review_folder):
        os.makedirs(review_folder)
        
    # Setup the Excel Workbook
    wb, ws_details, ws_qc = setup_excel_workbook()
    
    # Red text for Failed QC status
    red_font = Font(color="9C0006", bold=True)

    print(f"Scanning folder: {input_folder}")
    
    # Loop through every file in the folder
    for filename in os.listdir(input_folder):
        if filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(input_folder, filename)
            print(f"\nProcessing: {filename}...")
            
            try:
                # 1. Extract Data
                extracted_data = extract_invoice_data(pdf_path)
                
                # 2. Calculate QC Metrics
                calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
                total_calculated = calculated_line_sum + extracted_data.tax_amount + extracted_data.shipping_handling
                variance = round(extracted_data.total_amount - total_calculated, 2)
                
                # 3. Check Confidence & Math
                confidence_fields = [
                    extracted_data.invoice_number_confidence, extracted_data.origin_confidence,
                    extracted_data.destination_confidence, extracted_data.total_amount_confidence
                ]
                for item in extracted_data.line_items:
                    if item.uom_confidence == "Low":
                        confidence_fields.append("Low")
                
                needs_review = (variance != 0.0) or ("Low" in confidence_fields)
                status = "FAIL - NEEDS REVIEW" if needs_review else "PASS"
                
                # 4. Write to 'Invoice Details' Sheet
                for item in extracted_data.line_items:
                    ws_details.append([
                        filename, item.page_number, extracted_data.vendor_name, 
                        extracted_data.invoice_number, extracted_data.invoice_number_confidence,
                        extracted_data.origin, extracted_data.origin_confidence,
                        extracted_data.destination, extracted_data.destination_confidence,
                        extracted_data.bill_to, extracted_data.ship_to, extracted_data.remit_to,
                        extracted_data.currency, item.material, item.description, 
                        item.quantity, item.uom, item.uom_confidence, item.unit_price, 
                        item.line_total, extracted_data.total_amount, extracted_data.total_amount_confidence
                    ])
                    
                # Append Shipping/Tax rows to Details if they exist
                if extracted_data.shipping_handling > 0:
                    ws_details.append([filename, "", extracted_data.vendor_name, extracted_data.invoice_number, "", "", "", "", "", "", "", "", extracted_data.currency, "", "Shipping/Handling", "", "", "", "", extracted_data.shipping_handling, extracted_data.total_amount, ""])
                if extracted_data.tax_amount > 0:
                    ws_details.append([filename, "", extracted_data.vendor_name, extracted_data.invoice_number, "", "", "", "", "", "", "", "", extracted_data.currency, "", "Tax", "", "", "", "", extracted_data.tax_amount, extracted_data.total_amount, ""])

                # 5. Write to 'QC Summary' Sheet
                qc_row = [
                    filename, extracted_data.invoice_number, status, 
                    extracted_data.total_amount, total_calculated, variance, extracted_data.currency
                ]
                ws_qc.append(qc_row)
                
                # Turn the Status text Red if it failed
                if needs_review:
                    ws_qc.cell(row=ws_qc.max_row, column=3).font = red_font
                
                # 6. Apply Auto-Routing
                if needs_review:
                    print(f"⚠️ FLAG: Copying {filename} to 'Needs_Manual_Review'")
                    shutil.copy(pdf_path, os.path.join(review_folder, filename))
                else:
                    print(f"✅ PASS: {filename} processed perfectly.")
                    
            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")
                
    # Auto-adjust column widths before saving
    for sheet in [ws_details, ws_qc]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 2, 50) # Cap width at 50

    # Save the final workbook
    wb.save(output_excel)
    print(f"\n🎉 All finished! Data saved to: {output_excel}")
