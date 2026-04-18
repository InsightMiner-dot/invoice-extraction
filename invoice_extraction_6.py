import fitz  # PyMuPDF
import base64
import instructor
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
from pydantic import BaseModel, Field
from typing import List, Optional
import time

from dotenv import load_dotenv
load_dotenv(override=True)

# ==============================================================
# 1. Define Data Schema (UPDATED FOR INLINE ORIGIN/DESTINATION)
# ==============================================================

class LineItem(BaseModel):
    page_number: Optional[int] = Field(None, description="Page number (starting at 1)")
    material: Optional[str] = Field(None, description="Material code, part number, or SKU")
    description: str = Field(description="Name or description of the item. If a tax or fee (e.g., 'Federal Excise Tax', 'Franchise Tax') is printed as a row INSIDE the main table, extract it here as a regular line item.")
    quantity: Optional[float] = Field(None, description="Number of items SHIPPED. STRICT RULE: If you see 'QTY B/O' (Backordered) alongside 'QTY SHP', ONLY extract the Shipped amount. Do not extract backordered quantities here.")
    uom: Optional[str] = Field(None, description="Unit of Measure (e.g., EA, LBS, KG). Look for headers like 'UOM' or 'Bin'.")
    uom_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    unit_price: Optional[float] = Field(None, description="Price of a single unit")
    line_total: Optional[float] = Field(None, description="Total cost for this specific line item. Look for 'Amount', 'Total', 'Extended Price', or 'Extd Price'. STRICT RULE: If the column is blank (e.g., due to a backordered item), leave this null or 0.0. Do not guess.")
    
    # ---> THE FIX: Line-Level Origin & Destination <---
    line_origin: Optional[str] = Field(None, description="ONLY extract if the Origin/Ship-From address is explicitly listed per-item INSIDE the table row.")
    line_destination: Optional[str] = Field(None, description="ONLY extract if the Destination/Ship-To address is explicitly listed per-item INSIDE the table row.")

class TaxItem(BaseModel):
    tax_name: str = Field(description="The exact printed name of the tax (e.g., 'GST/HST', 'TPS/TVH', 'QST').")
    tax_amount: float = Field(description="The amount for this specific tax.")

class FeeItem(BaseModel):
    fee_name: str = Field(description="The exact printed name of the fee (e.g., 'SHOP Supplies', 'Environmental Fee', 'Pallet Charge').")
    fee_amount: float = Field(description="The amount for this specific fee.")

class InvoiceData(BaseModel):
    vendor_name: str = Field(description="Name of the company issuing the invoice")
    vendor_address: Optional[str] = Field(None, description="The FULL complete address of the vendor including street, city, state/province, and postal code. Do not extract partial addresses.")
    bill_to: Optional[str] = Field(None, description="The FULL complete 'Bill To' or 'Sold To' address including street, city, state/province, and postal code.")
    remit_to: Optional[str] = Field(None, description="The FULL complete 'Remit To' address including street, city, state/province, and postal code.")
    
    origin: Optional[str] = Field(None, description="The FULL origin physical address for the overall invoice. STRICT GUARDRAIL: ONLY extract this if it is explicitly labeled with tags like 'Ship From', 'Origin', 'From', or 'Pickup'. Do NOT extract short alphanumeric codes or tank numbers.")
    origin_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    destination: Optional[str] = Field(None, description="The FULL destination physical address for the overall invoice ('Ship To', 'To', 'Deliver To', 'Consignee'). STRICT RULE: Do NOT extract short alphanumeric facility codes or building numbers.")
    destination_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    invoice_number: Optional[str] = Field(None, description="Unique invoice number. STRICT RULE: Only extract values explicitly labeled as 'Invoice Number' or 'Invoice #'. Do NOT extract Ticket No, Reference, Order ID, or Statement numbers.")
    invoice_number_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    date: Optional[str] = Field(None, description="Date the invoice was issued")
    currency: Optional[str] = Field(None, description="3-letter currency code")
    
    subtotal: Optional[float] = Field(None, description="The subtotal amount before taxes and shipping are added.")
    
    taxes: List[TaxItem] = Field(default_factory=list, description="STRICT RULE: Extract individual taxes ONLY from the summary block at the bottom of the invoice. CRITICAL GUARDRAIL: Before adding a tax here, verify you did not already extract it in `line_items`. No duplicates allowed. If a tax is already a line item, skip it here.")
    
    additional_fees: List[FeeItem] = Field(default_factory=list, description="STRICT RULE: ONLY extract fees from the summary block at the bottom. CRITICAL GUARDRAIL: Before adding a fee here, verify you did not already extract it in `line_items`. No duplicates allowed.")
    
    shipping_name: Optional[str] = Field(None, description="The exact printed name of the shipping charge (e.g., 'Freight', 'Handling', 'Delivery Fee').")
    shipping_handling: Optional[float] = Field(0.0, description="STRICT RULE: ONLY extract this if it appears in the final summary block at the bottom of the invoice, AFTER the main line items and subtotal.")
    
    total_amount: float = Field(description="Final total amount charged on the invoice")
    total_amount_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    line_items: List[LineItem] = Field(description="List of all individual items purchased. CRITICAL GUARDRAIL: You must extract EVERY SINGLE ROW from the invoice table. Do NOT skip, summarize, or abbreviate rows. You must capture 100% of the items to ensure accounting math is perfectly accurate.")

class InvoiceDocument(BaseModel):
    invoices: List[InvoiceData] = Field(description="List of distinct invoices. STRICT PAGING RULE: 1) If an invoice table extends across multiple pages but shares the SAME Invoice Number, MERGE all line items into ONE invoice record. 2) If the Invoice Number CHANGES on a new page, split it into a NEW separate invoice record in this list.")

# ==============================================================
# 2. Setup Azure OpenAI Client
# ==============================================================

azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
azure_api_key = os.getenv("AZURE_OPENAI_KEY")
azure_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")
azure_api_version = os.getenv("AZURE_OPENAI_API_VERSION")

client = instructor.from_openai(
    AzureOpenAI(azure_endpoint=azure_endpoint, api_key=azure_api_key, api_version=azure_api_version)
)

# ==============================================================
# 3. PDF to Image Conversion & Extraction
# ==============================================================

def pdf_to_base64_images(pdf_path: str, max_pages: int = 15) -> List[str]:
    doc = fitz.open(pdf_path)
    base64_images = []
    
    total_pages = len(doc)
    if total_pages > max_pages:
        print(f"   [INFO] PDF is {total_pages} pages long. Truncating to first {max_pages} pages to prevent API timeout.")
        
    pages_to_process = min(total_pages, max_pages)
    
    for page_num in range(pages_to_process):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150) 
        base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
    return base64_images

def extract_invoice_data(pdf_path: str) -> InvoiceDocument:
    base64_images = pdf_to_base64_images(pdf_path)
    content_array = [{"type": "text", "text": "Extract data from this multi-page document. Pay close attention to invoice numbers. Evaluate confidence ('High', 'Medium', 'Low'). Note page numbers."}]
    
    for img_base64 in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}})

    system_prompt = "You are an expert accountant processing a document that may contain multiple distinct invoices. STRICT PAGING RULES: 1) If the same invoice number continues across multiple pages, combine all line items, taxes, and totals into a SINGLE invoice record. 2) If you see a NEW invoice number, start a NEW invoice record. CRITICAL RULE AGAINST DUPLICATES: Never extract the same tax or fee twice. ANTI-LAZINESS RULE: DO NOT BE LAZY. You must extract every single line item row by row. Skipping the middle of a table, abbreviating, or summarizing items is a CRITICAL FAILURE. Your extracted total for EACH invoice must mathematically equal the calculated sum of unique items for that invoice."

    return client.chat.completions.create(
        model=azure_deployment, 
        response_model=InvoiceDocument, 
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": content_array}
        ]
    )

# ==============================================================
# 4. Excel Setup Utility 
# ==============================================================

def setup_excel_workbook():
    wb = openpyxl.Workbook()
    
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Vendor Address", "Bill To", "Remit To",
        "Origin", "Destination", "Invoice Number", "Date", "Currency", 
        "Material", "Description", "Quantity", "UOM", "Unit Price", "Line Total", "Subtotal", "Invoice Total",
        "Inv# Conf", "Origin Conf", "Dest Conf", "UOM Conf", "Total Conf",
        "Status", "Reason for Review"
    ]
    ws_details.append(details_headers)
    
    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "File Name", "Vendor Name", "Invoice Number", "Status", "Reason for Review", 
        "Extracted Total", "Calculated Sum (Lines+Taxes+Ship+Fees)", "Variance"
    ]
    ws_qc.append(qc_headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    return wb, ws_details, ws_qc

# ==============================================================
# 5. Main Execution & Error Diagnostics
# ==============================================================

if __name__ == "__main__":
    start_time = time.time()
    
    input_folder = r"C:\path\to\your\invoices"
    output_excel = os.path.join(input_folder, "Extracted_Invoices_Master.xlsx")
        
    wb, ws_details, ws_qc = setup_excel_workbook()
    red_font = Font(color="9C0006", bold=True)

    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    total_pdfs = len(pdf_files)
    
    print(f"Scanning folder: {input_folder}")
    print(f"Found {total_pdfs} PDF(s) to process.\n")
    
    success_count = 0
    error_count = 0
    
    for current_index, filename in enumerate(pdf_files, start=1):
        pdf_path = os.path.join(input_folder, filename)
        
        print(f"\n[{current_index}/{total_pdfs}] Processing File: {filename}...")
        
        try:
            extracted_document = extract_invoice_data(pdf_path)
            
            if not extracted_document.invoices:
                print(f"   ⚠️ FLAG: No invoices detected in {filename}")
                ws_qc.append([filename, "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found in PDF", "", "", ""])
                ws_qc.cell(row=ws_qc.max_row, column=4).font = red_font
                continue
            
            print(f"   [INFO] Found {len(extracted_document.invoices)} distinct invoice(s) inside this PDF.")

            for extracted_data in extracted_document.invoices:
                
                calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
                safe_ship = extracted_data.shipping_handling if extracted_data.shipping_handling is not None else 0.0
                safe_total_tax = sum(tax.tax_amount for tax in extracted_data.taxes if tax.tax_amount is not None)
                safe_fees = sum(fee.fee_amount for fee in extracted_data.additional_fees if fee.fee_amount is not None)
                
                total_calculated = calculated_line_sum + safe_total_tax + safe_ship + safe_fees
                variance = round(extracted_data.total_amount - total_calculated, 2)
                
                review_reasons = []
                
                if variance != 0.0:
                    review_reasons.append(f"Math Variance of {variance}")
                    
                if extracted_data.invoice_number and extracted_data.invoice_number_confidence == "Low":
                    review_reasons.append("Low Conf: Invoice #")
                if extracted_data.origin and extracted_data.origin_confidence == "Low":
                    review_reasons.append("Low Conf: Origin")
                if extracted_data.destination and extracted_data.destination_confidence == "Low":
                    review_reasons.append("Low Conf: Destination")
                if extracted_data.total_amount_confidence == "Low":
                    review_reasons.append("Low Conf: Total Amount")
                
                if len(extracted_data.line_items) == 0:
                    review_reasons.append("Missing: 0 Line Items Found")
                    
                for item in extracted_data.line_items:
                    if item.uom and item.uom_confidence == "Low":
                        short_desc = item.description[:15] + "..." if item.description and len(item.description) > 15 else item.description
                        review_reasons.append(f"Low Conf: UOM on '{short_desc}'")
                
                needs_review = len(review_reasons) > 0
                status = "FAIL - NEEDS REVIEW" if needs_review else "PASS"
                reasons_string = " | ".join(review_reasons) if needs_review else "N/A"
                
                # ---> THE FIX: Fallback Logic added to the row creator <---
                def create_row(page_num, material, desc, qty, uom, uom_conf, price, line_total, line_orig, line_dest):
                    # Check line item first. If blank, fallback to the global invoice address.
                    final_origin = line_orig if line_orig else extracted_data.origin
                    final_dest = line_dest if line_dest else extracted_data.destination
                    
                    return [
                        filename, page_num, extracted_data.vendor_name, extracted_data.vendor_address,
                        extracted_data.bill_to, extracted_data.remit_to, 
                        final_origin, final_dest, extracted_data.invoice_number,
                        extracted_data.date, extracted_data.currency, 
                        material, desc, qty, uom, price, line_total, extracted_data.subtotal, extracted_data.total_amount,
                        extracted_data.invoice_number_confidence, extracted_data.origin_confidence, 
                        extracted_data.destination_confidence, uom_conf, extracted_data.total_amount_confidence,
                        status, reasons_string
                    ]

                if len(extracted_data.line_items) == 0:
                    ws_details.append(create_row(None, None, None, None, None, None, None, 0.0, None, None))
                else:
                    for item in extracted_data.line_items:
                        ws_details.append(create_row(
                            item.page_number, item.material, item.description, item.quantity, 
                            item.uom, item.uom_confidence, item.unit_price, item.line_total,
                            item.line_origin, item.line_destination # Pass the line-level addresses here
                        ))
                    
                if safe_ship > 0:
                    ship_label = extracted_data.shipping_name if extracted_data.shipping_name else "Shipping/Handling"
                    ws_details.append(create_row(None, None, ship_label, None, None, None, None, safe_ship, None, None))
                    
                for tax in extracted_data.taxes:
                    if tax.tax_amount is not None and tax.tax_amount > 0:
                        ws_details.append(create_row(None, None, tax.tax_name, None, None, None, None, tax.tax_amount, None, None))
                        
                for fee in extracted_data.additional_fees:
                    if fee.fee_amount is not None and fee.fee_amount > 0:
                        ws_details.append(create_row(None, None, fee.fee_name, None, None, None, None, fee.fee_amount, None, None))

                ws_qc.append([
                    filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                    status, reasons_string, extracted_data.total_amount, total_calculated, variance
                ])
                
                if needs_review:
                    ws_qc.cell(row=ws_qc.max_row, column=4).font = red_font
                    print(f"   ⚠️ FLAG: Invoice #{extracted_data.invoice_number} failed due to: {reasons_string}")
                else:
                    print(f"   ✅ PASS: Invoice #{extracted_data.invoice_number} processed perfectly.")
                    
            success_count += 1

        except Exception as e:
            error_msg = f"System/API Crash: {str(e)}"
            print(f"   ❌ CRITICAL ERROR: {error_msg}")
            
            error_row_details = [filename] + [""] * 23 + ["FAIL - CRITICAL ERROR", error_msg]
            ws_details.append(error_row_details)
            
            error_row_qc = [filename, "N/A", "N/A", "FAIL - CRITICAL ERROR", error_msg, "", "", ""]
            ws_qc.append(error_row_qc)
            ws_qc.cell(row=ws_qc.max_row, column=4).font = red_font
            
            error_count += 1
        
        print("   Waiting 2 seconds to prevent API rate limits...")
        time.sleep(2)
                
    for sheet in [ws_details, ws_qc]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 2, 50) 

    wb.save(output_excel)
    
    end_time = time.time()
    total_seconds = end_time - start_time
    minutes = int(total_seconds // 60)
    seconds = total_seconds % 60
    
    print("\n" + "="*50)
    print(" BATCH PROCESSING COMPLETE ")
    print("="*50)
    print(f"📄 Total PDFs Scanned:         {total_pdfs}")
    print(f"✅ Total PDFs Extracted:       {success_count} (Data written to Excel)")
    if error_count > 0:
        print(f"❌ Critical API Errors:        {error_count} (PDFs that crashed)")
    print("-" * 50)
    print(f"⏱️ Total Execution Time:       {minutes} min {seconds:.1f} sec")
    print(f"💾 Data saved to: {output_excel}")
    print("="*50 + "\n")
