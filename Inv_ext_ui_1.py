import streamlit as st
import fitz  # PyMuPDF
import base64
import instructor
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Tuple
import time
import io
from datetime import datetime
from dotenv import load_dotenv
import pandas as pd
import plotly.express as px

# Load environment variables from the backend
load_dotenv(override=True)

AZURE_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_API_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
AZURE_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")

# ==============================================================
# 0. Database Setup & Query Functions
# ==============================================================

AUDIT_FOLDER = "audit"
DB_PATH = os.path.join(AUDIT_FOLDER, "qc_master_database.sqlite")

def init_db():
    os.makedirs(AUDIT_FOLDER, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS qc_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            extraction_date TEXT,
            extraction_time TEXT,
            file_name TEXT,
            vendor_name TEXT,
            invoice_number TEXT,
            origin TEXT,
            destination TEXT,
            status TEXT,
            reason_for_review TEXT,
            extracted_total REAL,
            calculated_sum REAL,
            variance REAL,
            processing_time REAL,
            page_count INTEGER
        )
    ''')
    
    # Safe migration: Add new columns if updating from an older version of the DB
    try:
        cursor.execute("ALTER TABLE qc_audit ADD COLUMN processing_time REAL")
        cursor.execute("ALTER TABLE qc_audit ADD COLUMN page_count INTEGER")
    except sqlite3.OperationalError:
        pass 
        
    conn.commit()
    conn.close()

def insert_audit_record(record: tuple):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO qc_audit (
            extraction_date, extraction_time, file_name, vendor_name, 
            invoice_number, origin, destination, status, reason_for_review, 
            extracted_total, calculated_sum, variance, processing_time, page_count
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', record)
    conn.commit()
    conn.close()

@st.cache_data(ttl=3600)
def fetch_audit_data() -> pd.DataFrame:
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM qc_audit", conn)
    conn.close()
    return df

init_db()

# ==============================================================
# 1. Define Data Schema (Your Exact Core Logic + Custom Fields)
# ==============================================================

class LineItem(BaseModel):
    page_number: Optional[int] = Field(None, description="Page number (starting at 1)")
    material: Optional[str] = Field(None, description="Material code, part number, or SKU")
    description: str = Field(description="Name or description of the item. If a tax or fee (e.g., 'Federal Excise Tax', 'Franchise Tax') is printed as a row INSIDE the main table, extract it here as a regular line item.")
    quantity: Optional[float] = Field(None, description="Number of items SHIPPED. STRICT RULE: If you see 'QTY B/O' (Backordered) alongside 'QTY SHP', ONLY extract the Shipped amount. Do not extract backordered quantities here.")
    uom: Optional[str] = Field(None, description="Unit of Measure (e.g., EA, LBS, KG). Look for headers like 'UOM' or 'Bin'.")
    uom_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    unit_price: Optional[float] = Field(None, description="Price of a single unit")
    line_total: Optional[float] = Field(None, description="Total cost for this specific line item. Look for 'Amount', 'Total', 'Extended Price', or 'Extd Price'. STRICT RULE: If the column is blank (e.g., due to a backordered item), leave this null or 0.0. Do not guess.")

class TaxItem(BaseModel):
    tax_name: str = Field(description="The exact printed name of the tax (e.g., 'GST/HST', 'TPS/TVH', 'QST').")
    tax_amount: float = Field(description="The amount for this specific tax.")

class FeeItem(BaseModel):
    fee_name: str = Field(description="The exact printed name of the fee (e.g., 'SHOP Supplies', 'Environmental Fee', 'Pallet Charge').")
    fee_amount: float = Field(description="The amount for this specific fee.")

class InvoiceData(BaseModel):
    vendor_name: str = Field(description="Name of the company issuing the invoice")
    vendor_address: Optional[str] = Field(None, description="The FULL complete address of the vendor including street, city, state/province, and postal code. Do not extract partial addresses.")
    bill_to: Optional[str] = Field(None, description="The FULL complete 'Bill To' or 'Sold To' address including street, city, state/province, and postal code.")
    remit_to: Optional[str] = Field(None, description="The FULL complete 'Remit To' address including street, city, state/province, and postal code.")
    
    origin: Optional[str] = Field(None, description="The FULL origin physical address. STRICT GUARDRAIL: ONLY extract this if it is explicitly labeled with tags like 'Ship From', 'Origin', 'From', or 'Pickup'. If these specific tags are missing, or if it is just a random secondary address, you MUST return null. Do NOT extract short alphanumeric codes or tank numbers.")
    origin_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    destination: Optional[str] = Field(None, description="The FULL destination physical address ('Ship To', 'To', 'Deliver To', 'Consignee'). STRICT RULE: Do NOT extract short alphanumeric facility codes or building numbers. If a full physical address is not present, leave null.")
    destination_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    invoice_number: Optional[str] = Field(None, description="Unique invoice number. STRICT RULE: Only extract values explicitly labeled as 'Invoice Number' or 'Invoice #'. Do NOT extract Ticket No, Reference, Order ID, or Statement numbers.")
    invoice_number_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    date: Optional[str] = Field(None, description="Date the invoice was issued")
    currency: Optional[str] = Field(None, description="3-letter currency code")
    
    subtotal: Optional[float] = Field(None, description="The subtotal amount before taxes and shipping are added.")
    
    taxes: List[TaxItem] = Field(default_factory=list, description="STRICT RULE: Extract individual taxes ONLY from the summary block at the bottom of the invoice. CRITICAL GUARDRAIL: Before adding a tax here, verify you did not already extract it in `line_items`. No duplicates allowed. If a tax is already a line item, skip it here.")
    
    additional_fees: List[FeeItem] = Field(default_factory=list, description="STRICT RULE: ONLY extract fees from the summary block at the bottom. CRITICAL GUARDRAIL: Before adding a fee here, verify you did not already extract it in `line_items`. No duplicates allowed.")
    
    shipping_name: Optional[str] = Field(None, description="The exact printed name of the shipping charge (e.g., 'Freight', 'Handling', 'Delivery Fee').")
    shipping_handling: Optional[float] = Field(0.0, description="STRICT RULE: ONLY extract this if it appears in the final summary block at the bottom of the invoice, AFTER the main line items and subtotal.")
    
    total_amount: float = Field(description="Final total amount charged on the invoice")
    total_amount_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    custom_fields: Dict[str, Optional[str]] = Field(default_factory=dict, description="Extract any custom fields requested by the user. The keys MUST match exactly.")
    
    line_items: List[LineItem] = Field(description="List of all individual items purchased")

class InvoiceDocument(BaseModel):
    invoices: List[InvoiceData] = Field(description="List of distinct invoices. STRICT PAGING RULE: 1) If an invoice table extends across multiple pages but shares the SAME Invoice Number, MERGE all line items into ONE invoice record. 2) If the Invoice Number CHANGES on a new page, split it into a NEW separate invoice record in this list.")

# ==============================================================
# 2. PDF to Image Conversion & Extraction
# ==============================================================

def pdf_to_base64_images(file_bytes: bytes, max_pages: int, dpi: int) -> Tuple[List[str], int]:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    base64_images = []
    total_pages = len(doc)
    if total_pages > max_pages:
        st.warning(f"PDF is {total_pages} pages long. Truncating to first {max_pages} pages based on configuration.")
    pages_to_process = min(total_pages, max_pages)
    for page_num in range(pages_to_process):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=dpi) 
        base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
    return base64_images, total_pages

def extract_invoice_data(client, deployment: str, file_bytes: bytes, custom_cols: List[str], max_pages: int, dpi: int) -> Tuple[InvoiceDocument, int]:
    base64_images, total_pages = pdf_to_base64_images(file_bytes, max_pages, dpi)
    
    system_prompt = "You are an expert accountant processing a document that may contain multiple distinct invoices. STRICT PAGING RULES: 1) If the same invoice number continues across multiple pages, combine all line items, taxes, and totals into a SINGLE invoice record. 2) If you see a NEW invoice number, start a NEW invoice record. CRITICAL RULE AGAINST DUPLICATES: Never extract the same tax or fee twice. Your extracted total for EACH invoice must mathematically equal the calculated sum of unique items for that invoice."
    
    if custom_cols:
        system_prompt += f"\n\nSTRICT RULE: The user has requested custom data extraction. You MUST also search the invoice for the following fields and place them in the 'custom_fields' dictionary: {', '.join(custom_cols)}. If a field is not found, leave its value as null."

    content_array = [{"type": "text", "text": "Extract data from this multi-page document. Pay close attention to invoice numbers. Evaluate confidence ('High', 'Medium', 'Low'). Note page numbers."}]
    for img_base64 in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}})

    response = client.chat.completions.create(
        model=deployment, 
        response_model=InvoiceDocument, 
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": content_array}]
    )
    return response, total_pages

def setup_excel_workbook(custom_cols: List[str]):
    wb = openpyxl.Workbook()
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Vendor Address", "Bill To", "Remit To",
        "Origin", "Destination", "Invoice Number", "Date", "Currency", 
        "Material", "Description", "Quantity", "UOM", "Unit Price", "Line Total", "Subtotal", "Invoice Total",
        "Inv# Conf", "Origin Conf", "Dest Conf", "UOM Conf", "Total Conf", "Status", "Reason for Review"
    ]
    if custom_cols: details_headers.extend(custom_cols)
    ws_details.append(details_headers)
    
    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "File Name", "Vendor Name", "Invoice Number", "Origin", "Destination", "Status", "Reason for Review", 
        "Extracted Total", "Calculated Sum", "Variance"
    ]
    ws_qc.append(qc_headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
    return wb, ws_details, ws_qc

# ==============================================================
# 3. Streamlit UI & Dashboard
# ==============================================================

st.set_page_config(page_title="AI Invoice Intelligence", page_icon="🧾", layout="wide")

st.title("🧾 AI Invoice Intelligence Platform")

# Sidebar
with st.sidebar:
    with st.expander("⚙️ Processing Configuration", expanded=True):
        config_max_pages = st.number_input("Max Pages per PDF", min_value=1, max_value=100, value=15, help="Truncate documents longer than this to save API limits.")
        config_dpi = st.slider("Render Resolution (DPI)", min_value=72, max_value=600, value=300, step=72, help="Lower DPI runs faster. Higher DPI reads small text better.")

    st.divider()

    st.header("➕ Custom Extraction")
    custom_columns_input = st.text_area("Custom Columns (Comma-separated)", placeholder="e.g., PO Number, Cost Center")
    custom_columns_list = [col.strip() for col in custom_columns_input.split(",")] if custom_columns_input.strip() else []
    
    st.divider()
    
    st.header("📄 Upload Files")
    uploaded_files = st.file_uploader("Upload PDF Documents", type=["pdf"], accept_multiple_files=True)

# Tabs
tab_extract, tab_dashboard = st.tabs(["⚙️ Extraction Suite", "📊 Analytics Dashboard"])

# ---------------------------------------------------------
# TAB 1: EXTRACTION SUITE
# ---------------------------------------------------------
with tab_extract:
    st.write("Upload invoices via the sidebar and click below to process them. Results will appear here instantly.")
    
    if st.button("🚀 Start Extraction", type="primary"):
        if not uploaded_files:
            st.error("Please upload at least one PDF file from the sidebar.")
        elif not all([AZURE_ENDPOINT, AZURE_API_KEY, AZURE_DEPLOYMENT]):
            st.error("⚠️ Azure credentials missing. Check your `.env` file.")
        else:
            try:
                client = instructor.from_openai(AzureOpenAI(azure_endpoint=AZURE_ENDPOINT, api_key=AZURE_API_KEY, api_version=AZURE_API_VERSION))
            except Exception as e:
                st.error(f"Failed to initialize AI client: {e}")
                st.stop()

            wb, ws_details, ws_qc = setup_excel_workbook(custom_columns_list)
            red_font = Font(color="9C0006", bold=True)
            
            success_count, error_count = 0, 0
            
            progress_bar = st.progress(0, text="Initializing processing sequence...")
            status_text = st.empty()
            
            current_run_summary = []
            current_run_details = []

            start_time_batch = time.time()
            total_files = len(uploaded_files)

            for idx, file in enumerate(uploaded_files):
                filename = file.name
                now = datetime.now()
                current_date, current_time = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")

                status_text.markdown(f"**Extracting ({idx+1}/{total_files}):** `{filename}`...")
                
                file_start_time = time.time()
                
                try:
                    file_bytes = file.read()
                    extracted_document, total_pages = extract_invoice_data(client, AZURE_DEPLOYMENT, file_bytes, custom_columns_list, config_max_pages, config_dpi)
                    
                    file_proc_time = round(time.time() - file_start_time, 2)
                    
                    if not extracted_document.invoices:
                        ws_qc.append([filename, "N/A", "N/A", "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found in PDF", "", "", ""])
                        insert_audit_record((current_date, current_time, filename, "N/A", "N/A", "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found", 0.0, 0.0, 0.0, file_proc_time, total_pages))
                        current_run_summary.append({"File Name": filename, "Vendor Name": "N/A", "Invoice #": "N/A", "Origin": "N/A", "Destination": "N/A", "Status": "FAIL", "Reason": "No Invoices Found in PDF"})
                        continue
                    
                    for extracted_data in extracted_document.invoices:
                        calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
                        safe_ship = extracted_data.shipping_handling or 0.0
                        safe_total_tax = sum(tax.tax_amount for tax in extracted_data.taxes if tax.tax_amount is not None)
                        safe_fees = sum(fee.fee_amount for fee in extracted_data.additional_fees if fee.fee_amount is not None)
                        
                        total_calculated = calculated_line_sum + safe_total_tax + safe_ship + safe_fees
                        variance = round(extracted_data.total_amount - total_calculated, 2)
                        
                        review_reasons = []
                        if variance != 0.0: review_reasons.append(f"Math Variance of {variance}")
                        if extracted_data.invoice_number_confidence == "Low": review_reasons.append("Low Conf: Invoice #")
                        if extracted_data.origin_confidence == "Low": review_reasons.append("Low Conf: Origin")
                        if extracted_data.destination_confidence == "Low": review_reasons.append("Low Conf: Destination")
                        if extracted_data.total_amount_confidence == "Low": review_reasons.append("Low Conf: Total Amount")
                        if len(extracted_data.line_items) == 0: review_reasons.append("Missing: 0 Line Items Found")
                        
                        for item in extracted_data.line_items:
                            if item.uom_confidence == "Low":
                                short_desc = item.description[:15] + "..." if item.description and len(item.description) > 15 else item.description
                                review_reasons.append(f"Low Conf: UOM on '{short_desc}'")
                        
                        needs_review = len(review_reasons) > 0
                        status = "FAIL - NEEDS REVIEW" if needs_review else "PASS"
                        reasons_string = " | ".join(review_reasons) if needs_review else "N/A"
                        
                        def create_row(page_num, material, desc, qty, uom, uom_conf, price, line_total):
                            base_row = [
                                filename, page_num, extracted_data.vendor_name, extracted_data.vendor_address,
                                extracted_data.bill_to, extracted_data.remit_to, 
                                extracted_data.origin, extracted_data.destination, extracted_data.invoice_number,
                                extracted_data.date, extracted_data.currency, 
                                material, desc, qty, uom, price, line_total, extracted_data.subtotal, extracted_data.total_amount,
                                extracted_data.invoice_number_confidence, extracted_data.origin_confidence, 
                                extracted_data.destination_confidence, uom_conf, extracted_data.total_amount_confidence,
                                status, reasons_string
                            ]
                            custom_values = [extracted_data.custom_fields.get(col, "Not Found") for col in custom_columns_list]
                            return base_row + custom_values

                        def append_ui_detail_row(page_num, material, desc, qty, uom, unit_price, line_total):
                            current_run_details.append({
                                "File Name": filename,
                                "Vendor": extracted_data.vendor_name,
                                "Invoice #": extracted_data.invoice_number,
                                "Page #": page_num,
                                "Material/Fee": material,
                                "Description": desc,
                                "Qty": qty,
                                "UOM": uom,
                                "Price": unit_price,
                                "Line Total": line_total
                            })

                        if len(extracted_data.line_items) == 0:
                            ws_details.append(create_row(None, None, None, None, None, None, None, 0.0))
                            append_ui_detail_row(None, None, "NO ITEMS FOUND", None, None, None, 0.0)
                        else:
                            for item in extracted_data.line_items:
                                ws_details.append(create_row(item.page_number, item.material, item.description, item.quantity, item.uom, item.uom_confidence, item.unit_price, item.line_total))
                                append_ui_detail_row(item.page_number, item.material, item.description, item.quantity, item.uom, item.unit_price, item.line_total)
                                
                        if safe_ship > 0: 
                            ws_details.append(create_row(None, None, extracted_data.shipping_name or "Shipping", None, None, None, None, safe_ship))
                            append_ui_detail_row(None, "SHIPPING", extracted_data.shipping_name or "Shipping", None, None, None, safe_ship)
                            
                        for tax in extracted_data.taxes:
                            if tax.tax_amount is not None and tax.tax_amount > 0: 
                                ws_details.append(create_row(None, None, tax.tax_name, None, None, None, None, tax.tax_amount))
                                append_ui_detail_row(None, "TAX", tax.tax_name, None, None, None, tax.tax_amount)
                                
                        for fee in extracted_data.additional_fees:
                            if fee.fee_amount is not None and fee.fee_amount > 0: 
                                ws_details.append(create_row(None, None, fee.fee_name, None, None, None, None, fee.fee_amount))
                                append_ui_detail_row(None, "FEE", fee.fee_name, None, None, None, fee.fee_amount)
                        
                        ws_qc.append([
                            filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                            extracted_data.origin, extracted_data.destination,
                            status, reasons_string, extracted_data.total_amount, total_calculated, variance
                        ])
                        
                        insert_audit_record((
                            current_date, current_time, filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                            str(extracted_data.origin), str(extracted_data.destination), status, reasons_string, 
                            extracted_data.total_amount, total_calculated, variance, file_proc_time, total_pages
                        ))

                        current_run_summary.append({
                            "File Name": filename,
                            "Vendor Name": extracted_data.vendor_name,
                            "Invoice #": extracted_data.invoice_number,
                            "Origin": extracted_data.origin or "Missing",
                            "Destination": extracted_data.destination or "Missing",
                            "Variance": f"${variance:,.2f}",
                            "Status": "✅ PASS" if status == "PASS" else "⚠️ FAIL",
                            "Proc Time": f"{file_proc_time}s"
                        })
                        success_count += 1
                        
                except Exception as e:
                    file_proc_time = round(time.time() - file_start_time, 2)
                    insert_audit_record((current_date, current_time, filename, "N/A", "N/A", "N/A", "N/A", "FAIL - CRITICAL ERROR", str(e), 0.0, 0.0, 0.0, file_proc_time, 0))
                    current_run_summary.append({"File Name": filename, "Vendor Name": "ERROR", "Invoice #": "ERROR", "Origin": "ERROR", "Destination": "ERROR", "Status": "❌ ERROR", "Reason": "API/System Crash", "Proc Time": f"{file_proc_time}s"})
                    error_count += 1
                
                if idx < total_files - 1: time.sleep(3) 
                
                files_processed = idx + 1
                elapsed_time = time.time() - start_time_batch
                avg_time_per_file = elapsed_time / files_processed
                remaining_files = total_files - files_processed
                est_time_remaining = avg_time_per_file * remaining_files
                
                mins, secs = divmod(int(est_time_remaining), 60)
                eta_string = f" | ETA: {mins}m {secs}s" if est_time_remaining > 0 else ""
                
                progress_bar.progress(files_processed / total_files, text=f"Processing {files_processed}/{total_files} documents{eta_string}")

            fetch_audit_data.clear()

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            progress_bar.empty()
            status_text.empty()
            st.success(f"🎉 Batch Processing Complete! Invoices Extracted: {success_count} | Errors: {error_count}")
            
            st.subheader("📝 Line Item Details (Original)")
            st.write("Granular breakdown of every extracted item, tax, and fee.")
            st.dataframe(pd.DataFrame(current_run_details), use_container_width=True, hide_index=True)

            st.divider()

            st.subheader("🛡️ QC Summary")
            st.write("High-level financial audit and mathematical validation.")
            st.dataframe(pd.DataFrame(current_run_summary), use_container_width=True, hide_index=True)

            st.download_button(
                label="📥 Download Detailed Excel Report",
                data=excel_buffer,
                file_name=f"Extracted_Invoices_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

# ---------------------------------------------------------
# TAB 2: ANALYTICS DASHBOARD
# ---------------------------------------------------------
with tab_dashboard:
    st.header("📈 Master Extraction Analytics")
    df_audit = fetch_audit_data()
    
    if df_audit.empty:
        st.info("📭 No data available. Process some invoices in the Extraction Suite to generate analytics.")
    else:
        missing_flags = ['N/A', 'None', '', 'null', 'None']
        df_audit['is_origin_missing'] = df_audit['origin'].isin(missing_flags) | df_audit['origin'].isna()
        df_audit['is_dest_missing'] = df_audit['destination'].isin(missing_flags) | df_audit['destination'].isna()

        st.subheader("Financial & Accuracy Overview")
        col1, col2, col3, col4 = st.columns(4)
        total_invoices = len(df_audit)
        passed_invoices = len(df_audit[df_audit['status'] == 'PASS'])
        accuracy_rate = (passed_invoices / total_invoices) * 100 if total_invoices > 0 else 0
        total_value = df_audit['extracted_total'].sum()
        total_variance = df_audit['variance'].abs().sum()

        col1.metric("Total Invoices Processed", f"{total_invoices:,}")
        col2.metric("Extraction Accuracy", f"{accuracy_rate:.1f}%")
        col3.metric("Total Extracted Value", f"${total_value:,.2f}")
        col4.metric("Total Discrepancy (Variance)", f"${total_variance:,.2f}", delta_color="inverse")

        st.subheader("System Performance & Logistics")
        loc1, loc2, loc3, loc4 = st.columns(4)
        
        missing_origin_count = df_audit['is_origin_missing'].sum()
        missing_dest_count = df_audit['is_dest_missing'].sum()
        origin_miss_rate = (missing_origin_count / total_invoices) * 100 if total_invoices > 0 else 0
        dest_miss_rate = (missing_dest_count / total_invoices) * 100 if total_invoices > 0 else 0
        
        avg_proc_time = df_audit['processing_time'].mean() if 'processing_time' in df_audit.columns else 0.0
        total_pages = df_audit['page_count'].sum() if 'page_count' in df_audit.columns else 0
        
        loc1.metric("⚠️ Missing Origins", f"{missing_origin_count:,}", f"{origin_miss_rate:.1f}% of total", delta_color="inverse")
        loc2.metric("⚠️ Missing Destinations", f"{missing_dest_count:,}", f"{dest_miss_rate:.1f}% of total", delta_color="inverse")
        loc3.metric("Total Pages Read", f"{total_pages:,}")
        loc4.metric("Avg. Processing Time", f"{avg_proc_time:.1f}s", "Per Document")

        st.divider()

        c1, c2 = st.columns(2)

        with c1:
            st.subheader("Status Distribution")
            df_audit['Clean_Status'] = df_audit['status'].apply(lambda x: 'PASS' if x == 'PASS' else 'FAIL')
            status_counts = df_audit['Clean_Status'].value_counts().reset_index()
            status_counts.columns = ['Status', 'Count']
            
            fig_pie = px.pie(
                status_counts, names='Status', values='Count', hole=0.4, color='Status',
                color_discrete_map={'PASS':'#2ecc71', 'FAIL':'#e74c3c'}
            )
            st.plotly_chart(fig_pie, use_container_width=True)

        with c2:
            st.subheader("Top Vendors by Invoice Volume")
            vendor_counts = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]['vendor_name'].value_counts().reset_index().head(10)
            vendor_counts.columns = ['Vendor', 'Invoice Count']
            
            fig_bar = px.bar(
                vendor_counts, x='Invoice Count', y='Vendor', orientation='h',
                color='Invoice Count', color_continuous_scale='Blues'
            )
            fig_bar.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_bar, use_container_width=True)

        st.divider()
        
        c3, c4 = st.columns(2)
        
        with c3:
            st.subheader("Top Vendors by Financial Value")
            vendor_value = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('vendor_name')['extracted_total'].sum().reset_index()
            vendor_value = vendor_value.sort_values(by='extracted_total', ascending=False).head(10)
            vendor_value.columns = ['Vendor', 'Total Value ($)']
            
            fig_val = px.bar(
                vendor_value, x='Total Value ($)', y='Vendor', orientation='h',
                color='Total Value ($)', color_continuous_scale='Greens'
            )
            fig_val.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_val, use_container_width=True)

        with c4:
            st.subheader("⚠️ Top Reasons for Review")
            df_fails = df_audit[df_audit['Clean_Status'] == 'FAIL']
            if not df_fails.empty:
                reasons_series = df_fails['reason_for_review'].str.split(" | ").explode()
                reason_counts = reasons_series.value_counts().reset_index().head(10)
                reason_counts.columns = ['Reason', 'Frequency']
                fig_err = px.bar(reason_counts, x='Reason', y='Frequency', color='Frequency', color_continuous_scale='Reds')
                st.plotly_chart(fig_err, use_container_width=True)
            else:
                st.success("No failed extractions to analyze! Great job.")

        st.divider()

        st.subheader("📍 Vendor Logistics Map (Origin & Destination)")
        st.write("Consolidated view of all unique shipping locations utilized by each vendor.")
        
        df_routes = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]

        def get_unique_clean(series):
            vals = [str(x) for x in series if pd.notna(x) and str(x) not in missing_flags]
            return " | ".join(sorted(list(set(vals)))) if vals else "⚠️ Missing"

        vendor_routes = df_routes.groupby('vendor_name').agg({
            'id': 'count',
            'origin': get_unique_clean,
            'destination': get_unique_clean
        }).reset_index().rename(columns={
            'vendor_name': 'Vendor Name',
            'id': 'Invoice Count', 
            'origin': 'Unique Origins', 
            'destination': 'Unique Destinations'
        }).sort_values(by='Invoice Count', ascending=False)
        
        st.dataframe(vendor_routes, use_container_width=True, hide_index=True)
