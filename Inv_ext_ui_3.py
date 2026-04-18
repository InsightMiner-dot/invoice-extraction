import streamlit as st
import fitz  # PyMuPDF
import base64
import instructor
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Tuple
import time
import io
from datetime import datetime
from dotenv import load_dotenv
import pandas as pd
import plotly.express as px

# Load environment variables from the backend
load_dotenv(override=True)

AZURE_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
AZURE_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")

# ==============================================================
# 0. Database Setup & Query Functions
# ==============================================================

AUDIT_FOLDER = "audit"
DB_PATH = os.path.join(AUDIT_FOLDER, "qc_master_database.sqlite")

def init_db():
    os.makedirs(AUDIT_FOLDER, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Main Extraction Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS qc_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            extraction_date TEXT,
            extraction_time TEXT,
            file_name TEXT,
            vendor_name TEXT,
            invoice_number TEXT,
            origin TEXT,
            destination TEXT,
            status TEXT,
            reason_for_review TEXT,
            extracted_total REAL,
            calculated_sum REAL,
            variance REAL,
            processing_time REAL,
            page_count INTEGER
        )
    ''')
    
    # Audit Log for Database Deletions
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS deletion_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            user_name TEXT,
            action_type TEXT,
            details TEXT
        )
    ''')
    
    try:
        cursor.execute("ALTER TABLE qc_audit ADD COLUMN processing_time REAL")
        cursor.execute("ALTER TABLE qc_audit ADD COLUMN page_count INTEGER")
    except sqlite3.OperationalError:
        pass 
        
    conn.commit()
    conn.close()

def insert_audit_record(record: tuple):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO qc_audit (
            extraction_date, extraction_time, file_name, vendor_name, 
            invoice_number, origin, destination, status, reason_for_review, 
            extracted_total, calculated_sum, variance, processing_time, page_count
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', record)
    conn.commit()
    conn.close()

@st.cache_data(ttl=3600)
def fetch_audit_data() -> pd.DataFrame:
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM qc_audit", conn)
    conn.close()
    return df

def remove_duplicates_db(user_name: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT COUNT(*) FROM qc_audit 
        WHERE id NOT IN (SELECT MIN(id) FROM qc_audit GROUP BY vendor_name, invoice_number)
        AND invoice_number NOT IN ('N/A', 'ERROR', '')
    ''')
    count = cursor.fetchone()[0]
    
    cursor.execute('''
        DELETE FROM qc_audit
        WHERE id NOT IN (
            SELECT MIN(id)
            FROM qc_audit
            GROUP BY vendor_name, invoice_number
        ) AND invoice_number NOT IN ('N/A', 'ERROR', '')
    ''')
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("INSERT INTO deletion_logs (timestamp, user_name, action_type, details) VALUES (?, ?, ?, ?)", 
                   (now, user_name, "REMOVE_DUPLICATES", f"Deleted {count} duplicate records based on Vendor+Invoice match."))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

def wipe_master_db(user_name: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM qc_audit")
    count = cursor.fetchone()[0]
    
    cursor.execute("DELETE FROM qc_audit")
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("INSERT INTO deletion_logs (timestamp, user_name, action_type, details) VALUES (?, ?, ?, ?)", 
                   (now, user_name, "WIPE_DATABASE", f"Wiped entire Master Database ({count} records deleted)."))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

init_db()

# ==============================================================
# 1. Define Data Schema
# ==============================================================

class LineItem(BaseModel):
    page_number: Optional[int] = Field(None, description="Page number (starting at 1)")
    material: Optional[str] = Field(None, description="Material code, part number, or SKU")
    description: str = Field(description="Name or description of the item. If a tax or fee (e.g., 'Federal Excise Tax', 'Franchise Tax') is printed as a row INSIDE the main table, extract it here as a regular line item.")
    quantity: Optional[float] = Field(None, description="Number of items SHIPPED. STRICT RULE: If you see 'QTY B/O' (Backordered) alongside 'QTY SHP', ONLY extract the Shipped amount. Do not extract backordered quantities here.")
    uom: Optional[str] = Field(None, description="Unit of Measure (e.g., EA, LBS, KG). Look for headers like 'UOM' or 'Bin'.")
    uom_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    unit_price: Optional[float] = Field(None, description="Price of a single unit")
    line_total: Optional[float] = Field(None, description="Total cost for this specific line item. Look for 'Amount', 'Total', 'Extended Price', or 'Extd Price'. STRICT RULE: If the column is blank (e.g., due to a backordered item), leave this null or 0.0. Do not guess.")
    
    line_origin: Optional[str] = Field(
        None, 
        description="CRITICAL ANTI-LAZINESS RULE: You MUST read the exact Origin/Ship-From address printed for THIS specific row. DO NOT copy or repeat the value from the row above. Every single row is unique. Look closely at the image for this exact line. If it is blank for this specific row, you MUST leave it null."
    )
    line_destination: Optional[str] = Field(
        None, 
        description="CRITICAL ANTI-LAZINESS RULE: You MUST read the exact Destination/Ship-To address printed for THIS specific row. DO NOT copy or repeat the value from the row above. Every single row is unique. Look closely at the image for this exact line. If it is blank for this specific row, you MUST leave it null."
    )

class TaxItem(BaseModel):
    tax_name: str = Field(description="The exact printed name of the tax (e.g., 'GST/HST', 'TPS/TVH', 'QST').")
    tax_amount: float = Field(description="The amount for this specific tax.")

class FeeItem(BaseModel):
    fee_name: str = Field(description="The exact printed name of the fee (e.g., 'SHOP Supplies', 'Environmental Fee', 'Pallet Charge').")
    fee_amount: float = Field(description="The amount for this specific fee.")

class InvoiceData(BaseModel):
    vendor_name: str = Field(description="Name of the company issuing the invoice")
    vendor_address: Optional[str] = Field(None, description="The FULL complete address of the vendor including street, city, state/province, and postal code. Do not extract partial addresses.")
    bill_to: Optional[str] = Field(None, description="The FULL complete 'Bill To' or 'Sold To' address including street, city, state/province, and postal code.")
    remit_to: Optional[str] = Field(None, description="The FULL complete 'Remit To' address including street, city, state/province, and postal code.")
    
    origin: Optional[str] = Field(None, description="The FULL origin physical address for the overall invoice. STRICT GUARDRAIL: ONLY extract this if it is explicitly labeled with tags like 'Ship From', 'Origin', 'From', or 'Pickup'. Do NOT extract short alphanumeric codes or tank numbers.")
    origin_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    destination: Optional[str] = Field(None, description="The FULL destination physical address for the overall invoice ('Ship To', 'To', 'Deliver To', 'Consignee'). STRICT RULE: Do NOT extract short alphanumeric facility codes or building numbers.")
    destination_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    invoice_number: Optional[str] = Field(None, description="Unique invoice number. STRICT RULE: Only extract values explicitly labeled as 'Invoice Number' or 'Invoice #'. Do NOT extract Ticket No, Reference, Order ID, or Statement numbers.")
    invoice_number_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    date: Optional[str] = Field(None, description="Date the invoice was issued")
    currency: Optional[str] = Field(None, description="3-letter currency code")
    
    subtotal: Optional[float] = Field(None, description="The subtotal amount before taxes and shipping are added.")
    
    taxes: List[TaxItem] = Field(default_factory=list, description="STRICT RULE: Extract individual taxes ONLY from the summary block at the bottom of the invoice. CRITICAL GUARDRAIL: Before adding a tax here, verify you did not already extract it in `line_items`. No duplicates allowed. If a tax is already a line item, skip it here.")
    
    additional_fees: List[FeeItem] = Field(default_factory=list, description="STRICT RULE: ONLY extract fees from the summary block at the bottom. CRITICAL GUARDRAIL: Before adding a fee here, verify you did not already extract it in `line_items`. No duplicates allowed.")
    
    shipping_name: Optional[str] = Field(None, description="The exact printed name of the shipping charge (e.g., 'Freight', 'Handling', 'Delivery Fee').")
    shipping_handling: Optional[float] = Field(0.0, description="STRICT RULE: ONLY extract this if it appears in the final summary block at the bottom of the invoice, AFTER the main line items and subtotal.")
    
    total_amount: float = Field(description="Final total amount charged on the invoice")
    total_amount_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    custom_fields: Dict[str, Optional[str]] = Field(default_factory=dict, description="Extract any custom fields requested by the user. The keys MUST match exactly.")
    
    line_items: List[LineItem] = Field(description="List of all individual items purchased. CRITICAL GUARDRAIL: You must extract EVERY SINGLE ROW from the invoice table. Do NOT skip, summarize, or abbreviate rows. You must capture 100% of the items to ensure accounting math is perfectly accurate.")

class InvoiceDocument(BaseModel):
    invoices: List[InvoiceData] = Field(description="List of distinct invoices. STRICT PAGING RULE: 1) If an invoice table extends across multiple pages but shares the SAME Invoice Number, MERGE all line items into ONE invoice record. 2) If the Invoice Number CHANGES on a new page, split it into a NEW separate invoice record in this list.")

# ==============================================================
# 2. PDF to Image Conversion & Extraction
# ==============================================================

def pdf_to_base64_images(file_bytes: bytes, max_pages: int, dpi: int) -> Tuple[List[str], int]:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    base64_images = []
    total_pages = len(doc)
    if total_pages > max_pages:
        st.warning(f"PDF is {total_pages} pages long. Truncating to first {max_pages} pages based on configuration.")
    pages_to_process = min(total_pages, max_pages)
    for page_num in range(pages_to_process):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=dpi) 
        base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
    return base64_images, total_pages

def extract_invoice_data(client, deployment: str, file_bytes: bytes, custom_fields_dict: Dict[str, str], standard_aliases_dict: Dict[str, str], max_pages: int, dpi: int) -> Tuple[InvoiceDocument, int]:
    base64_images, total_pages = pdf_to_base64_images(file_bytes, max_pages, dpi)
    
    system_prompt = "You are an expert accountant processing a document that may contain multiple distinct invoices. STRICT PAGING RULES: 1) If the same invoice number continues across multiple pages, combine all line items, taxes, and totals into a SINGLE invoice record. 2) If you see a NEW invoice number, start a NEW invoice record. CRITICAL RULE AGAINST DUPLICATES: Never extract the same tax or fee twice. ANTI-LAZINESS RULE: DO NOT BE LAZY. You must extract every single line item row by row. Skipping the middle of a table, abbreviating, or summarizing items is a CRITICAL FAILURE. Your extracted total for EACH invoice must mathematically equal the calculated sum of unique items for that invoice."
    
    if standard_aliases_dict:
        alias_str = "\n".join([f"- {k}: Also look for '{v}'" for k, v in standard_aliases_dict.items()])
        system_prompt += f"\n\nSTANDARD FIELD ALIASES:\nThe user has provided common aliases for standard schema fields. Use these to help locate the data:\n{alias_str}"
    
    if custom_fields_dict:
        rules_str = "\n".join([f"- Field Key: '{k}' | Definition & Aliases: {v}" for k, v in custom_fields_dict.items()])
        system_prompt += f"\n\nSTRICT RULE: The user has requested custom data extraction based on specific definitions and alias mappings. You MUST search the invoice for the following logic and place the results in the 'custom_fields' dictionary using the EXACT 'Field Key' provided:\n{rules_str}\nIf a field or its aliases are not found, leave its value as null."

    content_array = [{"type": "text", "text": "Extract data from this multi-page document. Pay close attention to invoice numbers. Evaluate confidence ('High', 'Medium', 'Low'). Note page numbers."}]
    for img_base64 in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}})

    response = client.chat.completions.create(
        model=deployment, 
        response_model=InvoiceDocument, 
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": content_array}]
    )
    return response, total_pages

def setup_excel_workbook(custom_cols: List[str]):
    wb = openpyxl.Workbook()
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Vendor Address", "Bill To", "Remit To",
        "Origin", "Destination", "Invoice Number", "Date", "Currency", 
        "Material", "Description", "Line Origin", "Line Destination", "Quantity", "UOM", "Unit Price", "Line Total", "Subtotal", "Invoice Total",
        "Inv# Conf", "Origin Conf", "Dest Conf", "UOM Conf", "Total Conf", "Status", "Reason for Review"
    ]
    if custom_cols: details_headers.extend(custom_cols)
    ws_details.append(details_headers)
    
    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "File Name", "Vendor Name", "Invoice Number", "Origin", "Destination", "Status", "Reason for Review", 
        "Extracted Total", "Calculated Sum", "Variance"
    ]
    ws_qc.append(qc_headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
    return wb, ws_details, ws_qc

# ==============================================================
# 3. Streamlit UI & Core Logic Orchestrator
# ==============================================================

if 'extraction_details' not in st.session_state:
    st.session_state.extraction_details = None
if 'extraction_summary' not in st.session_state:
    st.session_state.extraction_summary = None
if 'extraction_excel' not in st.session_state:
    st.session_state.extraction_excel = None
if 'start_processing' not in st.session_state:
    st.session_state.start_processing = False
if 'files_to_process' not in st.session_state:
    st.session_state.files_to_process = []

def run_extraction_process(files_list, custom_fields_dict, standard_aliases_dict, max_pages, dpi, sleep_time, prefix=""):
    client = instructor.from_openai(AzureOpenAI(azure_endpoint=AZURE_ENDPOINT, api_key=AZURE_API_KEY, api_version=AZURE_API_VERSION))
    
    custom_col_keys = list(custom_fields_dict.keys())
    wb, ws_details, ws_qc = setup_excel_workbook(custom_col_keys)
    red_font = Font(color="9C0006", bold=True)
    
    success_count, error_count = 0, 0
    progress_bar = st.progress(0, text=f"Initializing {prefix} processing sequence...")
    status_text = st.empty()
    
    current_run_summary = []
    current_run_details = []

    start_time_batch = time.time()
    total_files = len(files_list)

    for idx, file in enumerate(files_list):
        filename = file.name
        now = datetime.now()
        current_date, current_time = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")

        status_text.markdown(f"**Extracting ({idx+1}/{total_files}):** `{filename}`...")
        file_start_time = time.time()
        
        try:
            file.seek(0)
            file_bytes = file.read()
            extracted_document, total_pages = extract_invoice_data(client, AZURE_DEPLOYMENT, file_bytes, custom_fields_dict, standard_aliases_dict, max_pages, dpi)
            file_proc_time = round(time.time() - file_start_time, 2)
            
            if not extracted_document.invoices:
                ws_qc.append([filename, "N/A", "N/A", "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found in PDF", "", "", ""])
                insert_audit_record((current_date, current_time, filename, "N/A", "N/A", "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found", 0.0, 0.0, 0.0, file_proc_time, total_pages))
                current_run_summary.append({"File Name": filename, "Vendor Name": "N/A", "Invoice #": "N/A", "Origin": "N/A", "Destination": "N/A", "Status": "FAIL", "Reason": "No Invoices Found in PDF"})
                continue
            
            for extracted_data in extracted_document.invoices:
                calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
                safe_ship = extracted_data.shipping_handling or 0.0
                safe_total_tax = sum(tax.tax_amount for tax in extracted_data.taxes if tax.tax_amount is not None)
                safe_fees = sum(fee.fee_amount for fee in extracted_data.additional_fees if fee.fee_amount is not None)
                
                total_calculated = calculated_line_sum + safe_total_tax + safe_ship + safe_fees
                variance = round(extracted_data.total_amount - total_calculated, 2)
                
                review_reasons = []
                if variance != 0.0: review_reasons.append(f"Math Variance of {variance}")
                if extracted_data.invoice_number_confidence == "Low": review_reasons.append("Low Conf: Invoice #")
                if extracted_data.origin_confidence == "Low": review_reasons.append("Low Conf: Origin")
                if extracted_data.destination_confidence == "Low": review_reasons.append("Low Conf: Destination")
                if extracted_data.total_amount_confidence == "Low": review_reasons.append("Low Conf: Total Amount")
                if len(extracted_data.line_items) == 0: review_reasons.append("Missing: 0 Line Items Found")
                
                for item in extracted_data.line_items:
                    if item.uom_confidence == "Low":
                        short_desc = item.description[:15] + "..." if item.description and len(item.description) > 15 else item.description
                        review_reasons.append(f"Low Conf: UOM on '{short_desc}'")
                
                needs_review = len(review_reasons) > 0
                status = "FAIL - NEEDS REVIEW" if needs_review else "PASS"
                reasons_string = " | ".join(review_reasons) if needs_review else "N/A"
                
                def create_row_dict(page_num, material, desc, qty, uom, uom_conf, price, line_total, line_orig=None, line_dest=None):
                    final_origin = line_orig if line_orig else extracted_data.origin
                    final_dest = line_dest if line_dest else extracted_data.destination
                    
                    row_data = {
                        "File Name": filename, "Page #": page_num,
                        "Vendor Name": extracted_data.vendor_name, "Vendor Address": extracted_data.vendor_address,
                        "Bill To": extracted_data.bill_to, "Remit To": extracted_data.remit_to,
                        "Origin": final_origin, "Destination": final_dest, 
                        "Invoice Number": extracted_data.invoice_number, "Date": extracted_data.date, "Currency": extracted_data.currency,
                        "Material": material, "Description": desc, "Line Origin": line_orig, "Line Destination": line_dest,
                        "Quantity": qty, "UOM": uom, "Unit Price": price, "Line Total": line_total,
                        "Subtotal": extracted_data.subtotal, "Invoice Total": extracted_data.total_amount,
                        "Inv# Conf": extracted_data.invoice_number_confidence, "Origin Conf": extracted_data.origin_confidence,
                        "Dest Conf": extracted_data.destination_confidence, "UOM Conf": uom_conf, "Total Conf": extracted_data.total_amount_confidence,
                        "Status": status, "Reason for Review": reasons_string
                    }
                    for col in custom_col_keys:
                        row_data[col] = extracted_data.custom_fields.get(col, "Not Found")
                    return row_data

                if len(extracted_data.line_items) == 0:
                    row_dict = create_row_dict(None, None, "NO ITEMS FOUND", None, None, None, None, 0.0)
                    ws_details.append(list(row_dict.values()))
                    current_run_details.append(row_dict)
                else:
                    for item in extracted_data.line_items:
                        row_dict = create_row_dict(
                            item.page_number, item.material, item.description, item.quantity, 
                            item.uom, item.uom_confidence, item.unit_price, item.line_total, 
                            line_orig=item.line_origin, line_dest=item.line_destination
                        )
                        ws_details.append(list(row_dict.values()))
                        current_run_details.append(row_dict)
                        
                if safe_ship > 0: 
                    row_dict = create_row_dict(None, "SHIPPING", extracted_data.shipping_name or "Shipping", None, None, None, None, safe_ship)
                    ws_details.append(list(row_dict.values()))
                    current_run_details.append(row_dict)
                    
                for tax in extracted_data.taxes:
                    if tax.tax_amount is not None and tax.tax_amount > 0: 
                        row_dict = create_row_dict(None, "TAX", tax.tax_name, None, None, None, None, tax.tax_amount)
                        ws_details.append(list(row_dict.values()))
                        current_run_details.append(row_dict)
                        
                for fee in extracted_data.additional_fees:
                    if fee.fee_amount is not None and fee.fee_amount > 0: 
                        row_dict = create_row_dict(None, "FEE", fee.fee_name, None, None, None, None, fee.fee_amount)
                        ws_details.append(list(row_dict.values()))
                        current_run_details.append(row_dict)
                
                # Use first line origin for QC Summary table if available
                first_line_orig = extracted_data.line_items[0].line_origin if len(extracted_data.line_items) > 0 else None
                first_line_dest = extracted_data.line_items[0].line_destination if len(extracted_data.line_items) > 0 else None
                qc_origin = first_line_orig if first_line_orig else extracted_data.origin
                qc_dest = first_line_dest if first_line_dest else extracted_data.destination

                ws_qc.append([
                    filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                    qc_origin, qc_dest,
                    status, reasons_string, extracted_data.total_amount, total_calculated, variance
                ])
                
                insert_audit_record((
                    current_date, current_time, filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                    str(qc_origin), str(qc_dest), status, reasons_string, 
                    extracted_data.total_amount, total_calculated, variance, file_proc_time, total_pages
                ))

                current_run_summary.append({
                    "File Name": filename, "Vendor Name": extracted_data.vendor_name, "Invoice #": extracted_data.invoice_number,
                    "Origin": qc_origin or "Missing", "Destination": qc_dest or "Missing",
                    "Variance": f"${variance:,.2f}", "Status": "✅ PASS" if status == "PASS" else "⚠️ FAIL",
                    "Proc Time": f"{file_proc_time}s"
                })
                success_count += 1
                
        except Exception as e:
            file_proc_time = round(time.time() - file_start_time, 2)
            insert_audit_record((current_date, current_time, filename, "N/A", "N/A", "N/A", "N/A", "FAIL - CRITICAL ERROR", str(e), 0.0, 0.0, 0.0, file_proc_time, 0))
            current_run_summary.append({"File Name": filename, "Vendor Name": "ERROR", "Invoice #": "ERROR", "Origin": "ERROR", "Destination": "ERROR", "Status": "❌ ERROR", "Reason": "API/System Crash", "Proc Time": f"{file_proc_time}s"})
            error_count += 1
        
        if idx < total_files - 1: time.sleep(sleep_time) 
        
        files_processed = idx + 1
        elapsed_time = time.time() - start_time_batch
        avg_time_per_file = elapsed_time / files_processed
        remaining_files = total_files - files_processed
        mins, secs = divmod(int(avg_time_per_file * remaining_files), 60)
        eta_string = f" | ETA: {mins}m {secs}s" if (avg_time_per_file * remaining_files) > 0 else ""
        progress_bar.progress(files_processed / total_files, text=f"Processing {files_processed}/{total_files} documents{eta_string}")

    fetch_audit_data.clear()

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    st.session_state.extraction_details = current_run_details
    st.session_state.extraction_summary = current_run_summary
    st.session_state.extraction_excel = excel_buffer
    
    progress_bar.empty()
    status_text.empty()
    st.success(f"🎉 {prefix} Batch Processing Complete! Invoices Extracted: {success_count} | Errors: {error_count}")

# --- DB & UI Dialogs ---
@st.dialog("⚠️ Duplicate Files Detected")
def confirm_duplicates_dialog(duplicate_files, unique_files):
    st.warning(f"Found {len(duplicate_files)} file(s) that have already been processed and exist in the master database.")
    for f in duplicate_files:
        st.write(f"- `{f.name}`")
    st.write("Do you want to extract these again? (This will add duplicate entries to your audit logs).")
    
    col1, col2 = st.columns(2)
    if col1.button("Yes, Process All", type="primary"):
        st.session_state.files_to_process = unique_files + duplicate_files
        st.session_state.start_processing = True
        st.rerun()
    if col2.button("No, Unique Only"):
        if unique_files:
            st.session_state.files_to_process = unique_files
            st.session_state.start_processing = True
            st.rerun()
        else:
            st.warning("No unique files to process.")
            time.sleep(2)
            st.rerun()

@st.dialog("🗑️ Remove Duplicate Invoices")
def dialog_remove_duplicates():
    st.warning("This will permanently delete all duplicate records (keeping the oldest one) based on Vendor Name + Invoice Number.")
    user_name = st.text_input("Enter your name to authorize deletion:")
    if st.button("Confirm Deletion", type="primary"):
        if not user_name.strip():
            st.error("Name is required for the audit log.")
        else:
            remove_duplicates_db(user_name.strip())
            st.rerun()

@st.dialog("⚠️ Wipe Master Database")
def dialog_wipe_db():
    st.error("CRITICAL WARNING: This will permanently delete ALL extracted invoice data from the system.")
    user_name = st.text_input("Enter your name to authorize complete wipe:")
    if st.button("Confirm Complete Wipe", type="primary"):
        if not user_name.strip():
            st.error("Name is required for the audit log.")
        else:
            wipe_master_db(user_name.strip())
            st.rerun()

@st.dialog("📄 Fullscreen Document Viewer", width="large")
def view_fullscreen_pdf(file_bytes, file_name):
    st.markdown(f"### {file_name}")
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    
    html_content = '<div style="height: 75vh; overflow-y: scroll; border: 1px solid #ddd; padding: 20px; background-color: #525659; text-align: center;">'
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        img_b64 = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
        html_content += f'<img src="data:image/jpeg;base64,{img_b64}" style="width: 95%; max-width: 900px; margin-bottom: 20px; box-shadow: 0px 4px 10px rgba(0,0,0,0.5);"><br>'
    html_content += '</div>'
    
    st.markdown(html_content, unsafe_allow_html=True)


# ==============================================================
# 4. Streamlit App Layout
# ==============================================================

st.set_page_config(page_title="AI Invoice Intelligence", page_icon="🧾", layout="wide")
st.title("🧾 AI Invoice Intelligence Platform")

# Sidebar
with st.sidebar:
    with st.expander("⚙️ Processing Configuration", expanded=False):
        config_max_pages = st.number_input("Max Pages per PDF", min_value=1, max_value=100, value=15)
        config_dpi = st.slider("Render Resolution (DPI)", min_value=72, max_value=600, value=300, step=72)
        config_sleep_time = st.number_input("API Sleep Time (seconds)", min_value=0, max_value=60, value=3)

    st.divider()
    
    st.header("🎯 Extraction Rules")
    with st.expander("1. Standard Field Aliases", expanded=False):
        st.write("Add alternative names (comma-separated) for built-in fields to help the AI find them.")
        
        default_standard = pd.DataFrame({
            "Standard Field": [
                "invoice_number", "date", "vendor_name", "vendor_address", 
                "bill_to", "remit_to", "origin", "destination", "currency", 
                "subtotal", "shipping_name", "shipping_handling", "total_amount",
                "material", "description", "line_origin", "line_destination", "quantity", "uom", "unit_price", "line_total",
                "tax_name", "tax_amount", "fee_name", "fee_amount"
            ],
            "Aliases": [""] * 25
        })
        standard_df = st.data_editor(default_standard, disabled=["Standard Field"], use_container_width=True, hide_index=True)
        
        standard_aliases_dict = {}
        for _, row in standard_df.iterrows():
            if str(row["Aliases"]).strip() and str(row["Aliases"]) != "None":
                standard_aliases_dict[row["Standard Field"]] = str(row["Aliases"]).strip()
                
    with st.expander("2. New Custom Fields", expanded=False):
        st.write("Add entirely new columns to extract.")
        default_custom = pd.DataFrame({
            "Field Name": [""],
            "Description & Aliases": [""]
        })
        custom_df = st.data_editor(default_custom, num_rows="dynamic", use_container_width=True, hide_index=True)
        
        custom_fields_dict = {}
        for _, row in custom_df.iterrows():
            name = str(row["Field Name"]).strip()
            desc = str(row["Description & Aliases"]).strip()
            if name and name != "None":
                custom_fields_dict[name] = desc
    
    st.divider()
    st.header("📄 Upload Files")
    uploaded_files = st.file_uploader("Upload PDF Documents", type=["pdf"], accept_multiple_files=True)

# Tabs
tab_extract, tab_viewer, tab_analytics, tab_system = st.tabs([
    "⚙️ Extraction Suite", "📄 Document Viewer", "📊 Business Analytics", "🤖 System & Drift Analysis"
])

# ---------------------------------------------------------
# TAB 1: EXTRACTION SUITE
# ---------------------------------------------------------
with tab_extract:
    st.write("Upload invoices via the sidebar and click below to process them. Results will appear here instantly.")
    
    if st.button("🚀 Start Extraction", type="primary"):
        if not uploaded_files:
            st.error("Please upload at least one PDF file from the sidebar.")
        elif not all([AZURE_ENDPOINT, AZURE_API_KEY, AZURE_DEPLOYMENT]):
            st.error("⚠️ Azure credentials missing. Check your `.env` file.")
        else:
            df_audit_existing = fetch_audit_data()
            existing_filenames = df_audit_existing['file_name'].unique().tolist() if not df_audit_existing.empty else []
            
            unique_files = [f for f in uploaded_files if f.name not in existing_filenames]
            duplicate_files = [f for f in uploaded_files if f.name in existing_filenames]
            
            if duplicate_files:
                confirm_duplicates_dialog(duplicate_files, unique_files)
            else:
                st.session_state.files_to_process = unique_files
                st.session_state.start_processing = True

    if st.session_state.start_processing:
        run_extraction_process(
            st.session_state.files_to_process, 
            custom_fields_dict, 
            standard_aliases_dict, 
            config_max_pages, 
            config_dpi,
            config_sleep_time
        )
        st.session_state.start_processing = False
        st.rerun()

    if st.session_state.extraction_details is not None:
        st.subheader("📝 Line Item Details (Full Extraction)")
        st.dataframe(pd.DataFrame(st.session_state.extraction_details), use_container_width=True, hide_index=True)

        st.divider()

        st.subheader("🛡️ QC Summary")
        st.dataframe(pd.DataFrame(st.session_state.extraction_summary), use_container_width=True, hide_index=True)

        st.download_button(
            label="📥 Download Excel Report",
            data=st.session_state.extraction_excel,
            file_name=f"Extraction_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ---------------------------------------------------------
# TAB 2: DOCUMENT VIEWER
# ---------------------------------------------------------
with tab_viewer:
    st.header("📄 Document Viewer")
    
    if uploaded_files:
        st.write("Search and click '🔍 View Full Screen' to examine a document in detail.")
        
        search_query = st.text_input("🔍 Search by File Name", "").lower()
        filtered_files = [f for f in uploaded_files if search_query in f.name.lower()]
        
        if not filtered_files:
            st.warning("No files match your search query.")
        else:
            cols = st.columns(4)
            for idx, file in enumerate(filtered_files):
                col = cols[idx % 4]
                with col:
                    st.markdown(f"<div style='white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-size: 14px; margin-bottom: 5px;' title='{file.name}'><b>{file.name}</b></div>", unsafe_allow_html=True)
                    
                    file.seek(0)
                    doc = fitz.open(stream=file.read(), filetype="pdf")
                    if len(doc) > 0:
                        page = doc[0]
                        pix = page.get_pixmap(dpi=72)
                        img_b64 = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
                        st.markdown(f'<img src="data:image/jpeg;base64,{img_b64}" style="width: 100%; border: 1px solid #ccc; box-shadow: 2px 2px 5px rgba(0,0,0,0.1); margin-bottom: 10px;">', unsafe_allow_html=True)
                    
                    file.seek(0)
                    if st.button("🔍 View Full Screen", key=f"view_{idx}", use_container_width=True):
                        view_fullscreen_pdf(file.read(), file.name)
    else:
        st.info("Upload PDF documents in the sidebar to view them here.")

# ---------------------------------------------------------
# TAB 3: BUSINESS ANALYTICS
# ---------------------------------------------------------
with tab_analytics:
    st.header("📊 Business Analytics")
    df_audit = fetch_audit_data()
    
    if df_audit.empty:
        st.info("📭 No data available. Process some invoices in the Extraction Suite to generate analytics.")
    else:
        missing_flags = ['N/A', 'None', '', 'null', 'None']
        
        st.subheader("Invoice & Vendor Overview")
        
        valid_df = df_audit[~df_audit['invoice_number'].isin(['N/A', 'ERROR', '', None]) & ~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]
        duplicates_df = valid_df[valid_df.duplicated(subset=['vendor_name', 'invoice_number'], keep='first')]
        dup_count = len(duplicates_df)
        dup_spend = duplicates_df['extracted_total'].sum()

        col1, col2, col3, col4, col5 = st.columns(5)
        total_invoices = len(df_audit)
        total_vendors = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]['vendor_name'].nunique()
        total_spend = df_audit['extracted_total'].sum()

        col1.metric("Total Invoices", f"{total_invoices:,}")
        col2.metric("Total Unique Vendors", f"{total_vendors:,}")
        col3.metric("Total Spend", f"${total_spend / 1_000_000:,.2f}M")
        col4.metric("⚠️ Duplicates Found", f"{dup_count:,}", delta_color="inverse")
        col5.metric("⚠️ Duplicate Value", f"${dup_spend / 1_000_000:,.2f}M", delta_color="inverse")

        st.divider()
        
        df_audit['extraction_date_dt'] = pd.to_datetime(df_audit['extraction_date'])
        spend_time = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('extraction_date')['extracted_total'].sum().reset_index()
        spend_time['extracted_total_M'] = spend_time['extracted_total'] / 1_000_000 

        st.subheader("📈 Financial Spend Over Time ($M)")
        fig_spend = px.line(spend_time, x='extraction_date', y='extracted_total_M', markers=True)
        fig_spend.update_layout(xaxis_title="Date", yaxis_title="Total Extracted (Millions)")
        st.plotly_chart(fig_spend, use_container_width=True)

        st.divider()

        st.subheader("🚨 Routing Exception Report")
        st.write("Identifies specific invoices where the Origin or Destination is missing, allowing you to easily locate and review the source files in the Document Viewer.")
        
        routing_issues_df = df_audit[
            df_audit['origin'].isin(missing_flags) | df_audit['origin'].isna() |
            df_audit['destination'].isin(missing_flags) | df_audit['destination'].isna()
        ]
        
        if not routing_issues_df.empty:
            review_cols = ['file_name', 'vendor_name', 'invoice_number', 'origin', 'destination', 'reason_for_review']
            st.dataframe(routing_issues_df[review_cols], use_container_width=True, hide_index=True)
        else:
            st.success("Excellent! All extracted invoices have complete Origin and Destination data.")

        st.divider()

        st.subheader("🔍 Deep Vendor Insights (Reliability Matrix)")
        st.write("Evaluates vendors based on their ability to provide valid Origins, Destinations, and accurate mathematical totals.")
        
        vr_df = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].copy()
        vr_df['origin_pass'] = ~vr_df['origin'].isin(missing_flags)
        vr_df['dest_pass'] = ~vr_df['destination'].isin(missing_flags)
        vr_df['math_pass'] = vr_df['variance'] == 0.0
        
        vendor_rel = vr_df.groupby('vendor_name').agg(
            Volume=('id', 'count'),
            Origin_Extracted=('origin_pass', 'mean'),
            Dest_Extracted=('dest_pass', 'mean'),
            Math_Accuracy=('math_pass', 'mean')
        ).reset_index()
        
        for col in ['Origin_Extracted', 'Dest_Extracted', 'Math_Accuracy']:
            vendor_rel[col] = (vendor_rel[col] * 100).round(1).astype(str) + '%'
            
        vendor_rel = vendor_rel.sort_values('Volume', ascending=False)
        st.dataframe(vendor_rel, use_container_width=True, hide_index=True)

        st.divider()

        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Top Vendors by Invoice Volume")
            vendor_counts = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]['vendor_name'].value_counts().reset_index().head(10)
            vendor_counts.columns = ['Vendor', 'Invoice Count']
            # Removed color arguments, added text_auto for simple readability
            fig_bar = px.bar(vendor_counts, x='Invoice Count', y='Vendor', orientation='h', text_auto=True)
            fig_bar.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_bar, use_container_width=True)

        with c2:
            st.subheader("Top Vendors by Financial Value ($M)")
            vendor_value = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('vendor_name')['extracted_total'].sum().reset_index()
            vendor_value['extracted_total_M'] = vendor_value['extracted_total'] / 1_000_000
            vendor_value = vendor_value.sort_values(by='extracted_total_M', ascending=False).head(10)
            vendor_value.columns = ['Vendor', 'Total Value ($)', 'Total Value (Millions)']
            # Removed color arguments, added text_auto for simple readability
            fig_val = px.bar(vendor_value, x='Total Value (Millions)', y='Vendor', orientation='h', text_auto='.2s')
            fig_val.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_val, use_container_width=True)

        st.divider()

        c3, c4 = st.columns(2)
        with c3:
            st.subheader("Pareto Analysis (Vendor Spend)")
            vendor_spend = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('vendor_name')['extracted_total'].sum().sort_values(ascending=False).reset_index()
            vendor_spend['extracted_total_M'] = vendor_spend['extracted_total'] / 1_000_000
            vendor_spend['Cumulative %'] = (vendor_spend['extracted_total'].cumsum() / vendor_spend['extracted_total'].sum()) * 100
            
            fig_pareto = px.bar(vendor_spend.head(20), x='vendor_name', y='extracted_total_M', title='Top 20 Vendors by Spend ($M)')
            fig_pareto.add_scatter(x=vendor_spend['vendor_name'].head(20), y=vendor_spend['Cumulative %'].head(20), mode='lines+markers', yaxis='y2', name='Cumulative %', line=dict(color='red'))
            fig_pareto.update_layout(yaxis2=dict(overlaying='y', side='right', range=[0, 100], title='Cumulative Percentage (%)'), xaxis_title="Vendor", yaxis_title="Total Spend (Millions)")
            st.plotly_chart(fig_pareto, use_container_width=True)

        with c4:
            st.subheader("Invoice Value Distribution ($M)")
            valid_spend = df_audit[df_audit['extracted_total'] > 0].copy()
            valid_spend['extracted_total_M'] = valid_spend['extracted_total'] / 1_000_000
            
            if not valid_spend.empty:
                fig_hist = px.histogram(valid_spend, x='extracted_total_M', nbins=20, title='Distribution of Invoice Amounts')
                fig_hist.update_layout(xaxis_title="Invoice Amount (Millions)", yaxis_title="Count")
                st.plotly_chart(fig_hist, use_container_width=True)
            else:
                st.info("Not enough spend data for distribution.")

        st.divider()

        st.subheader("📍 Top Shipping Corridors")
        df_valid_routes = df_audit[~df_audit['origin'].isin(missing_flags) & ~df_audit['destination'].isin(missing_flags)].copy()
        
        if not df_valid_routes.empty:
            df_valid_routes['Corridor'] = df_valid_routes['origin'].astype(str) + " ➡️ " + df_valid_routes['destination'].astype(str)
            corridor_counts = df_valid_routes['Corridor'].value_counts().reset_index().head(10)
            corridor_counts.columns = ['Shipping Corridor', 'Volume']
            # Removed color argument, added text_auto
            fig_corridor = px.bar(corridor_counts, x='Volume', y='Shipping Corridor', orientation='h', text_auto=True)
            fig_corridor.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_corridor, use_container_width=True)
        else:
            st.info("Not enough valid Origin & Destination pairs to map corridors.")
            
        st.divider()
        
        st.subheader("📑 Duplicate Invoices Detected")
        if not duplicates_df.empty:
            duplicates_df['extracted_total_M'] = duplicates_df['extracted_total'].apply(lambda x: f"${x/1_000_000:,.2f}M")
            st.dataframe(duplicates_df[['vendor_name', 'invoice_number', 'file_name', 'extracted_total_M', 'extraction_date']], use_container_width=True, hide_index=True)
        else:
            st.success("No duplicate invoices detected in the master database.")

# ---------------------------------------------------------
# TAB 4: SYSTEM & DRIFT ANALYSIS
# ---------------------------------------------------------
with tab_system:
    st.header("🤖 System & Model Drift Analysis")
    df_audit = fetch_audit_data()
    
    if df_audit.empty:
        st.info("📭 No data available. Process some invoices to track system drift.")
    else:
        df_audit['extraction_date_dt'] = pd.to_datetime(df_audit['extraction_date'])
        df_audit['datetime'] = pd.to_datetime(df_audit['extraction_date'] + ' ' + df_audit['extraction_time'])
        df_audit = df_audit.sort_values('datetime')
        df_audit['is_critical_error'] = df_audit['status'].str.contains("CRITICAL ERROR", na=False)
        df_audit['Clean_Status'] = df_audit['status'].apply(lambda x: 'PASS' if x == 'PASS' else 'FAIL')
        df_audit['processing_time_min'] = df_audit['processing_time'] / 60.0

        st.subheader("System Performance & Mathematical Accuracy")
        sys1, sys2, sys3, sys4, sys5, sys6 = st.columns(6)
        
        total_invoices = len(df_audit)
        total_pages = df_audit['page_count'].sum() if 'page_count' in df_audit.columns else 0
        avg_proc_time_min = df_audit['processing_time_min'].mean() if 'processing_time_min' in df_audit.columns else 0.0
        time_per_page_min = (df_audit['processing_time_min'].sum() / total_pages) if total_pages > 0 else 0.0
        sys_error_rate = (df_audit['is_critical_error'].sum() / total_invoices) * 100 if total_invoices > 0 else 0
        
        passed_invoices = len(df_audit[df_audit['status'] == 'PASS'])
        accuracy_rate = (passed_invoices / total_invoices) * 100 if total_invoices > 0 else 0
        total_variance = df_audit['variance'].abs().sum()

        sys1.metric("Avg Proc Time", f"{avg_proc_time_min:.2f}m", "Per Doc")
        sys2.metric("Proc Speed", f"{time_per_page_min:.2f}m", "Per Page")
        sys3.metric("API Failure Rate", f"{sys_error_rate:.1f}%", delta_color="inverse")
        sys4.metric("Pages Indexed", f"{total_pages:,}")
        sys5.metric("Data Accuracy", f"{accuracy_rate:.1f}%")
        sys6.metric("Math Variance", f"${total_variance:,.2f}", delta_color="inverse")

        st.divider()

        s1, s2 = st.columns(2)
        with s1:
            st.subheader("Processing Speed Drift (Minutes)")
            fig_speed = px.line(df_audit, x='datetime', y='processing_time_min', markers=True, title='Document Processing Time')
            fig_speed.update_layout(xaxis_title="Time", yaxis_title="Minutes")
            st.plotly_chart(fig_speed, use_container_width=True)

        with s2:
            st.subheader("Processing Volume by Day of Week")
            df_audit['day_of_week'] = df_audit['extraction_date_dt'].dt.day_name()
            day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            dow_counts = df_audit['day_of_week'].value_counts().reindex(day_order).reset_index()
            dow_counts.columns = ['Day', 'Volume']
            # Removed color mapping, added text_auto
            fig_dow = px.bar(dow_counts, x='Day', y='Volume', text_auto=True)
            st.plotly_chart(fig_dow, use_container_width=True)
                
        st.divider()

        c3, c4 = st.columns(2)
        with c3:
            st.subheader("⚠️ Top Reasons for LLM Review")
            df_fails = df_audit[df_audit['Clean_Status'] == 'FAIL']
            if not df_fails.empty:
                reasons_series = df_fails['reason_for_review'].str.split(" | ").explode()
                reason_counts = reasons_series.value_counts().reset_index().head(10)
                reason_counts.columns = ['Reason', 'Frequency']
                # Removed color argument
                fig_err = px.bar(reason_counts, x='Reason', y='Frequency', text_auto=True)
                st.plotly_chart(fig_err, use_container_width=True)
            else:
                st.success("No failures to analyze.")

        with c4:
            st.subheader("💸 Absolute Variance Magnitude by Vendor")
            var_vendor = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('vendor_name')['variance'].apply(lambda x: x.abs().sum()).reset_index()
            var_vendor = var_vendor[var_vendor['variance'] > 0].sort_values(by='variance', ascending=False).head(10)
            if not var_vendor.empty:
                # Removed color mapping
                fig_var_vendor = px.bar(var_vendor, x='variance', y='vendor_name', orientation='h', text_auto=True)
                fig_var_vendor.update_layout(yaxis={'categoryorder':'total ascending'})
                st.plotly_chart(fig_var_vendor, use_container_width=True)
            else:
                st.success("No variance detected across any vendors.")
                
        st.divider()
            
        st.subheader("⚙️ Database Management")
        st.write("Use these controls to maintain system health and clear test data.")
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        with col_btn1:
            if st.button("🗑️ Remove Duplicates", use_container_width=True):
                dialog_remove_duplicates()
        with col_btn2:
            if st.button("⚠️ Wipe Database", type="secondary", use_container_width=True):
                dialog_wipe_db()
