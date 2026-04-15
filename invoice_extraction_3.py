import fitz  # PyMuPDF
import base64
import instructor
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
from pydantic import BaseModel, Field
from typing import List, Optional

# ==============================================================
# 1. Define Data Schema (MERGED SHIP TO & STRICT ORIGIN GUARDRAILS)
# ==============================================================

class LineItem(BaseModel):
    page_number: Optional[int] = Field(None, description="Page number (starting at 1)")
    material: Optional[str] = Field(None, description="Material code, part number, or SKU")
    description: str = Field(description="Name or description of the item")
    quantity: Optional[float] = Field(None, description="Number of items purchased")
    uom: Optional[str] = Field(None, description="Unit of Measure")
    uom_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    unit_price: Optional[float] = Field(None, description="Price of a single unit")
    line_total: float = Field(description="Total cost for this specific line item.")

class TaxItem(BaseModel):
    tax_name: str = Field(description="The exact printed name of the tax (e.g., 'GST/HST', 'TPS/TVH', 'QST').")
    tax_amount: float = Field(description="The amount for this specific tax.")

class InvoiceData(BaseModel):
    vendor_name: str = Field(description="Name of the company issuing the invoice")
    vendor_address: Optional[str] = Field(None, description="The FULL complete address of the vendor including street, city, state/province, and postal code. Do not extract partial addresses.")
    bill_to: Optional[str] = Field(None, description="The FULL complete 'Bill To' or 'Sold To' address including street, city, state/province, and postal code.")
    remit_to: Optional[str] = Field(None, description="The FULL complete 'Remit To' address including street, city, state/province, and postal code.")
    
    # ---> STRICT GUARDRAILS ADDED TO ORIGIN <---
    origin: Optional[str] = Field(None, description="The FULL origin physical address. STRICT GUARDRAIL: ONLY extract this if it is explicitly labeled with tags like 'Ship From', 'Origin', 'From', or 'Pickup'. If these specific tags are missing, or if it is just a random secondary address, you MUST return null. Do NOT extract short alphanumeric codes or tank numbers.")
    origin_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    # ---> DESTINATION NOW ACTS AS 'SHIP TO' AS WELL <---
    destination: Optional[str] = Field(None, description="The FULL destination physical address ('Ship To', 'To', 'Deliver To', 'Consignee'). STRICT RULE: Do NOT extract short alphanumeric facility codes or building numbers. If a full physical address is not present, leave null.")
    destination_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    invoice_number: Optional[str] = Field(None, description="Unique invoice ID or number")
    invoice_number_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    date: Optional[str] = Field(None, description="Date the invoice was issued")
    currency: Optional[str] = Field(None, description="3-letter currency code")
    
    subtotal: Optional[float] = Field(None, description="The subtotal amount before taxes and shipping are added.")
    
    taxes: List[TaxItem] = Field(default_factory=list, description="STRICT RULE: Extract ALL individual taxes (e.g., GST, PST, QST) listed in the summary block at the bottom of the invoice, AFTER the subtotal. If no taxes are listed, leave this list empty.")
    
    shipping_name: Optional[str] = Field(None, description="The exact printed name of the shipping charge (e.g., 'Freight', 'Handling', 'Delivery Fee').")
    shipping_handling: Optional[float] = Field(0.0, description="STRICT RULE: ONLY extract this if it appears in the final summary block at the bottom of the invoice, AFTER the main line items and subtotal.")
    
    total_amount: float = Field(description="Final total amount charged on the invoice")
    total_amount_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    line_items: List[LineItem] = Field(description="List of all individual items purchased")

# ==============================================================
# 2. Setup Azure OpenAI Client
# ==============================================================

azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
azure_api_key = os.getenv("AZURE_OPENAI_KEY")
azure_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")
azure_api_version = os.getenv("AZURE_OPENAI_API_VERSION")

client = instructor.from_openai(
    AzureOpenAI(azure_endpoint=azure_endpoint, api_key=azure_api_key, api_version=azure_api_version)
)

# ==============================================================
# 3. PDF to Image Conversion & Extraction
# ==============================================================

def pdf_to_base64_images(pdf_path: str) -> List[str]:
    doc = fitz.open(pdf_path)
    base64_images = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
    return base64_images

def extract_invoice_data(pdf_path: str) -> InvoiceData:
    base64_images = pdf_to_base64_images(pdf_path)
    content_array = [{"type": "text", "text": "Extract data from this multi-page invoice. Pay special attention to capturing every distinct tax listed below the subtotal. Evaluate confidence ('High', 'Medium', 'Low'). Note page numbers."}]
    
    for img_base64 in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}})

    return client.chat.completions.create(
        model=azure_deployment, 
        response_model=InvoiceData, 
        messages=[
            {"role": "system", "content": "You are an expert accountant. Extract data accurately, ensure all addresses extracted are full physical addresses and adhere to strict keyword rules, and provide honest confidence scores."},
            {"role": "user", "content": content_array}
        ]
    )

# ==============================================================
# 4. Excel Setup Utility 
# ==============================================================

def setup_excel_workbook():
    wb = openpyxl.Workbook()
    
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    
    # ---> REMOVED 'Ship To' FROM HEADERS <---
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Vendor Address", "Bill To", "Remit To",
        "Origin", "Destination", "Invoice Number", "Date", "Currency", 
        "Material", "Description", "Quantity", "UOM", "Unit Price", "Line Total", "Subtotal", "Invoice Total",
        "Inv# Conf", "Origin Conf", "Dest Conf", "UOM Conf", "Total Conf"
    ]
    ws_details.append(details_headers)
    
    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "File Name", "Invoice Number", "Status", "Reason for Review", "Extracted Total", 
        "Calculated Sum (Lines+Taxes+Ship)", "Variance", "Currency"
    ]
    ws_qc.append(qc_headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    return wb, ws_details, ws_qc

# ==============================================================
# 5. Main Execution & Error Diagnostics
# ==============================================================

if __name__ == "__main__":
    input_folder = r"C:\path\to\your\invoices"
    output_excel = os.path.join(input_folder, "Extracted_Invoices_Master.xlsx")
        
    wb, ws_details, ws_qc = setup_excel_workbook()
    red_font = Font(color="9C0006", bold=True)

    print(f"Scanning folder: {input_folder}")
    
    for filename in os.listdir(input_folder):
        if filename.lower().endswith(".pdf"):
            pdf_path = os.path.join(input_folder, filename)
            print(f"\nProcessing: {filename}...")
            
            try:
                extracted_data = extract_invoice_data(pdf_path)
                
                calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
                safe_ship = extracted_data.shipping_handling if extracted_data.shipping_handling is not None else 0.0
                safe_total_tax = sum(tax.tax_amount for tax in extracted_data.taxes if tax.tax_amount is not None)
                
                total_calculated = calculated_line_sum + safe_total_tax + safe_ship
                variance = round(extracted_data.total_amount - total_calculated, 2)
                
                # Reason Generation Logic
                review_reasons = []
                
                if variance != 0.0:
                    review_reasons.append(f"Math Variance of {variance}")
                if extracted_data.invoice_number_confidence == "Low":
                    review_reasons.append("Low Conf: Invoice #")
                if extracted_data.origin_confidence == "Low":
                    review_reasons.append("Low Conf: Origin")
                if extracted_data.destination_confidence == "Low":
                    review_reasons.append("Low Conf: Destination")
                if extracted_data.total_amount_confidence == "Low":
                    review_reasons.append("Low Conf: Total Amount")
                    
                for item in extracted_data.line_items:
                    if item.uom_confidence == "Low":
                        short_desc = item.description[:15] + "..." if item.description and len(item.description) > 15 else item.description
                        review_reasons.append(f"Low Conf: UOM on '{short_desc}'")
                
                needs_review = len(review_reasons) > 0
                status = "FAIL - NEEDS REVIEW" if needs_review else "PASS"
                reasons_string = " | ".join(review_reasons) if needs_review else "N/A"
                
                # Helper function (Removed ship_to)
                def create_row(page_num, material, desc, qty, uom, uom_conf, price, line_total):
                    return [
                        filename, page_num, extracted_data.vendor_name, extracted_data.vendor_address,
                        extracted_data.bill_to, extracted_data.remit_to,  # ship_to removed here
                        extracted_data.origin, extracted_data.destination, extracted_data.invoice_number,
                        extracted_data.date, extracted_data.currency, 
                        material, desc, qty, uom, price, line_total, extracted_data.subtotal, extracted_data.total_amount,
                        extracted_data.invoice_number_confidence, extracted_data.origin_confidence, 
                        extracted_data.destination_confidence, uom_conf, extracted_data.total_amount_confidence
                    ]

                # 1. Write actual line items
                for item in extracted_data.line_items:
                    ws_details.append(create_row(
                        item.page_number, item.material, item.description, item.quantity, 
                        item.uom, item.uom_confidence, item.unit_price, item.line_total
                    ))
                    
                # 2. Write Shipping row dynamically
                if safe_ship > 0:
                    ship_label = extracted_data.shipping_name if extracted_data.shipping_name else "Shipping/Handling"
                    ws_details.append(create_row(None, None, ship_label, None, None, None, None, safe_ship))
                    
                # 3. Write MULTIPLE Tax rows dynamically
                for tax in extracted_data.taxes:
                    if tax.tax_amount is not None and tax.tax_amount > 0:
                        ws_details.append(create_row(None, None, tax.tax_name, None, None, None, None, tax.tax_amount))

                # 4. Write QC Summary 
                ws_qc.append([
                    filename, extracted_data.invoice_number, status, reasons_string,
                    extracted_data.total_amount, total_calculated, variance, extracted_data.currency
                ])
                
                if needs_review:
                    ws_qc.cell(row=ws_qc.max_row, column=3).font = red_font
                    print(f"⚠️ FLAG: '{filename}' failed due to: {reasons_string}")
                else:
                    print(f"✅ PASS: {filename} processed perfectly.")
                    
            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")
                
    # Auto-adjust column widths
    for sheet in [ws_details, ws_qc]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 2, 50) 

    wb.save(output_excel)
    print(f"\n🎉 All finished! Data saved to: {output_excel}")
