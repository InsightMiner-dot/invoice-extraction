import fitz
import base64
import time
from datetime import datetime
import re
import json
import asyncio
import io
from PIL import Image
from fuzzywuzzy import process
from openai import AsyncAzureOpenAI
import instructor
import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

from core.schemas import InvoiceDocument
from core.database import insert_audit_record, load_master_suppliers, standardize_vendor, DB_PATH, AUDIT_FOLDER

load_dotenv(override=True)

client = instructor.from_openai(AsyncAzureOpenAI(
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview"),
    max_retries=5
))
DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")

# Threaded File to Base64 (Handles PDFs and Images)
def render_file_to_base64(file_bytes: bytes, filename: str, max_pages: int, dpi: int):
    base64_images = []
    raw_pdf_text_list = []
    total_pages = 1
    
    ext = filename.split('.')[-1].lower()
    
    if ext == 'pdf':
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        total_pages = len(doc)
        pages = min(total_pages, max_pages)
        for p in range(pages):
            page = doc[p]
            pix = page.get_pixmap(dpi=dpi)
            base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
            raw_pdf_text_list.append(page.get_text("text"))
            del pix
        doc.close()
        
    elif ext in ['png', 'jpg', 'jpeg', 'tif', 'tiff']:
        try:
            img = Image.open(io.BytesIO(file_bytes))
            if img.mode != 'RGB':
                img = img.convert('RGB')
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='JPEG', quality=85)
            base64_images.append(base64.b64encode(img_byte_arr.getvalue()).decode('utf-8'))
        except Exception as e:
            print(f"Failed to process image: {e}")

    return base64_images, total_pages, "".join(raw_pdf_text_list)

async def process_single_invoice(file_bytes: bytes, filename: str, batch_id: str, max_pages: int, dpi: int, aliases: dict, custom_fields: dict):
    start_time = time.time()
    
    base64_images, total_pages, raw_pdf_text = await asyncio.to_thread(
        render_file_to_base64, file_bytes, filename, max_pages, dpi
    )
    clean_raw_text = re.sub(r'\s+', '', raw_pdf_text).lower() if raw_pdf_text else ""

    system_prompt = "You are an expert accountant processing a document that may contain multiple distinct invoices. STRICT PAGING RULES: 1) If the same invoice number continues across multiple pages, combine all line items, taxes, and totals into a SINGLE invoice record. 2) If you see a NEW invoice number, start a NEW invoice record. CRITICAL RULE AGAINST DUPLICATES: Never extract the same tax or fee twice. ANTI-LAZINESS RULE: DO NOT BE LAZY. You must extract every single line item row by row."
    
    if aliases:
        alias_str = "\n".join([f"- {k}: Also look for '{v}'" for k, v in aliases.items()])
        system_prompt += f"\n\nSTANDARD FIELD ALIASES:\n{alias_str}"
        
    if custom_fields:
        rules_str = "\n".join([f"- Field Key: '{k}' | Definition & Aliases: {v}" for k, v in custom_fields.items()])
        system_prompt += f"\n\nSTRICT CUSTOM RULES:\n{rules_str}"

    content_array = [{"type": "text", "text": "Extract data from this multi-page document. Evaluate confidence ('High', 'Medium', 'Low'). Note page numbers."}]
    for img in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img}"}})

    try:
        extracted_doc = await client.chat.completions.create(
            model=DEPLOYMENT, response_model=InvoiceDocument, 
            messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": content_array}]
        )
    except Exception as e:
        return {"error": str(e), "filename": filename}

    file_proc_time = round(time.time() - start_time, 2)
    master_df = load_master_suppliers()
    master_list = master_df['Original_Supplier_Name'].tolist() if not master_df.empty else []
    
    now = datetime.now()
    current_date, current_time = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")

    summary_results = []
    detail_results = []

    for extracted_data in extracted_doc.invoices:
        raw_vendor = extracted_data.vendor_name
        
        if master_list and raw_vendor and raw_vendor not in ['N/A', 'ERROR', '']:
            match_result = process.extractOne(raw_vendor, master_list)
            orig_supp = match_result[0] if match_result and match_result[1] >= 95 else standardize_vendor(raw_vendor)
        else:
            orig_supp = standardize_vendor(raw_vendor)

        calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
        safe_ship = extracted_data.shipping_handling or 0.0
        safe_total_tax = sum(tax.tax_amount for tax in extracted_data.taxes if tax.tax_amount is not None)
        safe_fees = sum(fee.fee_amount for fee in extracted_data.additional_fees if fee.fee_amount is not None)
        
        total_calculated = calculated_line_sum + safe_total_tax + safe_ship + safe_fees
        variance = round(extracted_data.total_amount - total_calculated, 2)

        hallucination_alert = False
        if extracted_data.invoice_number and clean_raw_text:
            clean_inv = re.sub(r'\s+', '', extracted_data.invoice_number).lower()
            if clean_inv and clean_inv not in clean_raw_text:
                hallucination_alert = True

        review_reasons = []
        unique_uoms = set()
        
        if variance != 0.0: review_reasons.append(f"Math Variance of {variance}")
        if hallucination_alert: review_reasons.append("HALLUCINATION: Inv # not in doc")
        if extracted_data.invoice_number_confidence == "Low": review_reasons.append("Low Conf: Invoice #")
        if extracted_data.origin_confidence == "Low": review_reasons.append("Low Conf: Origin")
        if extracted_data.destination_confidence == "Low": review_reasons.append("Low Conf: Destination")
        if extracted_data.total_amount_confidence == "Low": review_reasons.append("Low Conf: Total Amount")
        if len(extracted_data.line_items) == 0: review_reasons.append("Missing: 0 Line Items Found")
        
        for item in extracted_data.line_items:
            if item.uom: unique_uoms.add(item.uom)
            if item.uom_confidence == "Low":
                short_desc = item.description[:15] + "..." if item.description and len(item.description) > 15 else item.description
                review_reasons.append(f"Low Conf: UOM on '{short_desc}'")

        needs_review = len(review_reasons) > 0
        
        # --- NEW STATUS TEXT ---
        status = "NEEDS HUMAN REVIEW" if needs_review else "PASS"
        reasons_string = " | ".join(review_reasons) if needs_review else "N/A"
        uom_string = ", ".join(list(unique_uoms)) if unique_uoms else "N/A"
        
        first_line_orig = extracted_data.line_items[0].line_origin if len(extracted_data.line_items) > 0 else None
        first_line_dest = extracted_data.line_items[0].line_destination if len(extracted_data.line_items) > 0 else None
        qc_origin = first_line_orig if first_line_orig else extracted_data.origin
        qc_dest = first_line_dest if first_line_dest else extracted_data.destination

        full_json_dump = extracted_data.model_dump_json()
        custom_json = json.dumps(extracted_data.custom_fields) if extracted_data.custom_fields else "{}"

        insert_audit_record((
            current_date, current_time, filename, raw_vendor, orig_supp, extracted_data.vendor_address, 
            extracted_data.bill_to, extracted_data.remit_to, extracted_data.invoice_number, extracted_data.date, 
            extracted_data.currency, str(qc_origin), None, None, str(qc_dest), None, None, 
            extracted_data.subtotal, safe_ship, extracted_data.total_amount, total_calculated, variance, 
            extracted_data.invoice_number_confidence, extracted_data.origin_confidence, 
            extracted_data.destination_confidence, extracted_data.total_amount_confidence, uom_string, 
            status, reasons_string, file_proc_time, total_pages, batch_id, custom_json, full_json_dump
        ))

        summary_results.append({
            "File Name": filename, "Vendor Name": raw_vendor, "Invoice #": extracted_data.invoice_number,
            "Variance": f"${variance:,.2f}", "Status": "✅ PASS" if status == "PASS" else "⚠️ NEEDS HUMAN REVIEW", 
            "Proc Time": f"{file_proc_time}s"
        })
        
        def create_ui_row(page_num, material, desc, qty, uom, price, line_total):
            row = {
                "File Name": filename, "Page #": page_num, "Original Supplier": orig_supp, 
                "Invoice Number": extracted_data.invoice_number, "Material": material,
                "Description": desc, "Qty": qty, "UOM": uom, "Price": price, "Line Total": line_total, 
                "Inv# Conf": extracted_data.invoice_number_confidence, "Total Conf": extracted_data.total_amount_confidence, 
                "Variance": variance, "Proc Time": f"{file_proc_time}s", 
                "Status": "✅ PASS" if status == "PASS" else "⚠️ NEEDS HUMAN REVIEW"
            }
            for col in list(custom_fields.keys()): row[col] = extracted_data.custom_fields.get(col, "Not Found")
            return row

        if len(extracted_data.line_items) == 0:
            detail_results.append(create_ui_row(None, None, "NO ITEMS FOUND", None, None, None, 0.0))
        else:
            for item in extracted_data.line_items:
                detail_results.append(create_ui_row(item.page_number, item.material, item.description, item.quantity, item.uom, item.unit_price, item.line_total))
        
        if safe_ship > 0:
            detail_results.append(create_ui_row(None, "SHIPPING", extracted_data.shipping_name or "Shipping", None, None, None, safe_ship))
        for tax in extracted_data.taxes:
            if tax.tax_amount: detail_results.append(create_ui_row(None, "TAX", tax.tax_name, None, None, None, tax.tax_amount))
        for fee in extracted_data.additional_fees:
            if fee.fee_amount: detail_results.append(create_ui_row(None, "FEE", fee.fee_name, None, None, None, fee.fee_amount))

    return {"summary": summary_results, "details": detail_results, "total_file_pages": total_pages}


def generate_excel_from_db(batch_id: str, custom_cols: list):
    os.makedirs(AUDIT_FOLDER, exist_ok=True)
    excel_path = os.path.join(AUDIT_FOLDER, f"{batch_id}.xlsx")
    
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(f"SELECT * FROM qc_audit WHERE batch_id = '{batch_id}'", conn)
    conn.close()

    if df.empty: return False

    wb = openpyxl.Workbook()
    
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Original Supplier Name", "Vendor Address", "Bill To", "Remit To",
        "Origin", "Destination", "Invoice Number", "Date", "Currency", 
        "Material", "Description", "Quantity", "UOM", "Unit Price", "Line Total", "Subtotal", "Invoice Total",
        "Inv# Conf", "Origin Conf", "Dest Conf", "UOM Conf", "Total Conf", "Status", "Reason for Review"
    ]
    excel_headers = details_headers.copy()
    if custom_cols: excel_headers.extend(custom_cols)
    ws_details.append(excel_headers)
    
    for _, row in df.iterrows():
        try: full_data = json.loads(row['line_items'])
        except: continue
            
        def create_excel_row(page_num, material, desc, qty, uom, uom_conf, price, line_total, line_orig, line_dest):
            final_orig = line_orig if line_orig else full_data.get('origin', '')
            final_dest = line_dest if line_dest else full_data.get('destination', '')
            base = [
                row['file_name'], page_num, row['vendor_name'], row['original_supplier_name'],
                full_data.get('vendor_address', ''), full_data.get('bill_to', ''), full_data.get('remit_to', ''),
                final_orig, final_dest, row['invoice_number'], row['invoice_date'], row['currency'],
                material, desc, qty, uom, price, line_total, row['subtotal'], row['extracted_total'],
                row['invoice_number_conf'], row['origin_conf'], row['destination_conf'], uom_conf,
                row['total_amount_conf'], row['status'], row['reason_for_review']
            ]
            cf_dict = full_data.get('custom_fields', {})
            for col in custom_cols: base.append(cf_dict.get(col, "Not Found"))
            return base

        line_items = full_data.get('line_items', [])
        if not line_items:
            ws_details.append(create_excel_row(None, None, "NO ITEMS FOUND", None, None, None, None, 0.0, None, None))
        else:
            for item in line_items:
                ws_details.append(create_excel_row(
                    item.get('page_number'), item.get('material'), item.get('description'), item.get('quantity'),
                    item.get('uom'), item.get('uom_confidence'), item.get('unit_price'), item.get('line_total'),
                    item.get('line_origin'), item.get('line_destination')
                ))
        
        safe_ship = full_data.get('shipping_handling')
        if safe_ship and float(safe_ship) > 0:
            ws_details.append(create_excel_row(None, "SHIPPING", full_data.get('shipping_name') or "Shipping", None, None, None, None, float(safe_ship), None, None))
            
        for tax in full_data.get('taxes', []):
            if tax.get('tax_amount'):
                ws_details.append(create_excel_row(None, "TAX", tax.get('tax_name'), None, None, None, None, tax.get('tax_amount'), None, None))
                
        for fee in full_data.get('additional_fees', []):
            if fee.get('fee_amount'):
                ws_details.append(create_excel_row(None, "FEE", fee.get('fee_name'), None, None, None, None, fee.get('fee_amount'), None, None))

    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "file_name", "vendor_name", "original_supplier_name", "invoice_number", 
        "origin", "destination", "status", "reason_for_review", 
        "extracted_total", "calculated_sum", "variance"
    ]
    ws_qc.append(["File Name", "Vendor Name", "Original Supplier Name", "Invoice Number", "Origin", "Destination", "Status", "Reason for Review", "Extracted Total", "Calculated Sum", "Variance"])
    
    for _, row in df.iterrows():
        ws_qc.append([row.get(h, '') for h in qc_headers])

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(excel_path)
    return True
