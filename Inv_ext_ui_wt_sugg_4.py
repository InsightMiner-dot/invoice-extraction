import streamlit as st
import fitz  # PyMuPDF
import base64
import instructor
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import AzureOpenAI
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Tuple
import time
import io
from datetime import datetime
import re
from dotenv import load_dotenv
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from wordcloud import WordCloud
import matplotlib.pyplot as plt

# Load environment variables from the backend
load_dotenv(override=True)

AZURE_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
AZURE_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")

# ==============================================================
# 0. Database & CSV Setup Functions
# ==============================================================

AUDIT_FOLDER = "audit"
DB_PATH = os.path.join(AUDIT_FOLDER, "qc_master_database.sqlite")
MASTER_CSV_PATH = os.path.join(AUDIT_FOLDER, "master_suppliers.csv")

def init_db():
    os.makedirs(AUDIT_FOLDER, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS qc_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            extraction_date TEXT,
            extraction_time TEXT,
            file_name TEXT,
            vendor_name TEXT,
            invoice_number TEXT,
            origin TEXT,
            destination TEXT,
            status TEXT,
            reason_for_review TEXT,
            extracted_total REAL,
            calculated_sum REAL,
            variance REAL,
            processing_time REAL,
            page_count INTEGER
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS deletion_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            user_name TEXT,
            action_type TEXT,
            details TEXT
        )
    ''')
    
    new_columns = [
        ("processing_time", "REAL"),
        ("page_count", "INTEGER"),
        ("clean_supplier", "TEXT"),
        ("suggested_origin", "TEXT"),
        ("final_origin", "TEXT"),
        ("suggested_destination", "TEXT"),
        ("final_destination", "TEXT"),
        ("batch_id", "TEXT")
    ]
    
    for col_name, col_type in new_columns:
        try:
            cursor.execute(f"ALTER TABLE qc_audit ADD COLUMN {col_name} {col_type}")
        except sqlite3.OperationalError:
            pass 
        
    conn.commit()
    conn.close()

# ---> NEW: Master Supplier CSV Functions <---
def load_master_suppliers() -> pd.DataFrame:
    if os.path.exists(MASTER_CSV_PATH):
        return pd.read_csv(MASTER_CSV_PATH)
    else:
        return pd.DataFrame(columns=["Raw_Vendor_Name", "Clean_Supplier_Name"])

def save_master_suppliers(df: pd.DataFrame):
    df.to_csv(MASTER_CSV_PATH, index=False)

def insert_audit_record(record: tuple):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO qc_audit (
            extraction_date, extraction_time, file_name, vendor_name, 
            invoice_number, origin, destination, status, reason_for_review, 
            extracted_total, calculated_sum, variance, processing_time, page_count, batch_id, clean_supplier
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', record)
    conn.commit()
    conn.close()

def update_audit_record_supplier(record_id: int, clean_supplier: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE qc_audit SET clean_supplier = ? WHERE id = ?", (clean_supplier, record_id))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

def update_audit_suggestions(record_id: int, sug_orig: str, sug_dest: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE qc_audit SET suggested_origin = ?, suggested_destination = ? WHERE id = ?", (sug_orig, sug_dest, record_id))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

def update_audit_final_routes(record_id: int, final_orig: str, final_dest: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("UPDATE qc_audit SET final_origin = ?, final_destination = ? WHERE id = ?", (final_orig, final_dest, record_id))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

@st.cache_data(ttl=3600)
def fetch_audit_data() -> pd.DataFrame:
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT * FROM qc_audit", conn)
    conn.close()
    return df

def remove_duplicates_db(user_name: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COUNT(*) FROM qc_audit 
        WHERE id NOT IN (SELECT MIN(id) FROM qc_audit GROUP BY vendor_name, invoice_number)
        AND invoice_number NOT IN ('N/A', 'ERROR', '')
    ''')
    count = cursor.fetchone()[0]
    cursor.execute('''
        DELETE FROM qc_audit
        WHERE id NOT IN (
            SELECT MIN(id)
            FROM qc_audit
            GROUP BY vendor_name, invoice_number
        ) AND invoice_number NOT IN ('N/A', 'ERROR', '')
    ''')
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("INSERT INTO deletion_logs (timestamp, user_name, action_type, details) VALUES (?, ?, ?, ?)", 
                   (now, user_name, "REMOVE_DUPLICATES", f"Deleted {count} duplicate records based on Vendor+Invoice match."))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

def wipe_master_db(user_name: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM qc_audit")
    count = cursor.fetchone()[0]
    cursor.execute("DELETE FROM qc_audit")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute("INSERT INTO deletion_logs (timestamp, user_name, action_type, details) VALUES (?, ?, ?, ?)", 
                   (now, user_name, "WIPE_DATABASE", f"Wiped entire Master Database ({count} records deleted)."))
    conn.commit()
    conn.close()
    fetch_audit_data.clear()

def standardize_vendor(name):
    if not isinstance(name, str) or name in ['N/A', 'ERROR', '']:
        return name
    name = name.upper()
    name = re.sub(r'[.,]', '', name)
    name = re.sub(r'\b(LLC|INC|LTD|CORP|CORPORATION|CO|COMPANY)\b', '', name)
    return re.sub(r'\s+', ' ', name).strip()

init_db()

# ==============================================================
# 1. Define Data Schema
# ==============================================================

class LineItem(BaseModel):
    page_number: Optional[int] = Field(None, description="Page number (starting at 1)")
    material: Optional[str] = Field(None, description="Material code, part number, or SKU")
    description: str = Field(description="Name or description of the item. If a tax or fee (e.g., 'Federal Excise Tax', 'Franchise Tax') is printed as a row INSIDE the main table, extract it here as a regular line item.")
    quantity: Optional[float] = Field(None, description="Number of items SHIPPED. STRICT RULE: If you see 'QTY B/O' (Backordered) alongside 'QTY SHP', ONLY extract the Shipped amount. Do not extract backordered quantities here.")
    uom: Optional[str] = Field(None, description="Unit of Measure (e.g., EA, LBS, KG). Look for headers like 'UOM' or 'Bin'.")
    uom_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    unit_price: Optional[float] = Field(None, description="Price of a single unit")
    line_total: Optional[float] = Field(None, description="Total cost for this specific line item. Look for 'Amount', 'Total', 'Extended Price', or 'Extd Price'. STRICT RULE: If the column is blank (e.g., due to a backordered item), leave this null or 0.0. Do not guess.")
    
    line_origin: Optional[str] = Field(None, description="CRITICAL ANTI-LAZINESS RULE: You MUST read the exact Origin/Ship-From address printed for THIS specific row. DO NOT copy or repeat the value from the row above. Every single row is unique. Look closely at the image for this exact line. If it is blank for this specific row, you MUST leave it null.")
    line_destination: Optional[str] = Field(None, description="CRITICAL ANTI-LAZINESS RULE: You MUST read the exact Destination/Ship-To address printed for THIS specific row. DO NOT copy or repeat the value from the row above. Every single row is unique. Look closely at the image for this exact line. If it is blank for this specific row, you MUST leave it null.")

class TaxItem(BaseModel):
    tax_name: str = Field(description="The exact printed name of the tax (e.g., 'GST/HST', 'TPS/TVH', 'QST').")
    tax_amount: float = Field(description="The amount for this specific tax.")

class FeeItem(BaseModel):
    fee_name: str = Field(description="The exact printed name of the fee (e.g., 'SHOP Supplies', 'Environmental Fee', 'Pallet Charge').")
    fee_amount: float = Field(description="The amount for this specific fee.")

class InvoiceData(BaseModel):
    vendor_name: str = Field(description="Name of the company issuing the invoice")
    vendor_address: Optional[str] = Field(None, description="The FULL complete address of the vendor including street, city, state/province, and postal code. Do not extract partial addresses.")
    bill_to: Optional[str] = Field(None, description="The FULL complete 'Bill To' or 'Sold To' address including street, city, state/province, and postal code.")
    remit_to: Optional[str] = Field(None, description="The FULL complete 'Remit To' address including street, city, state/province, and postal code.")
    
    origin: Optional[str] = Field(None, description="The FULL origin physical address for the overall invoice. STRICT GUARDRAIL: ONLY extract this if it is explicitly labeled with tags like 'Ship From', 'Origin', 'From', or 'Pickup'. Do NOT extract short alphanumeric codes or tank numbers.")
    origin_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    destination: Optional[str] = Field(None, description="The FULL destination physical address for the overall invoice ('Ship To', 'To', 'Deliver To', 'Consignee'). STRICT RULE: Do NOT extract short alphanumeric facility codes or building numbers.")
    destination_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    invoice_number: Optional[str] = Field(None, description="Unique invoice number. STRICT RULE: Only extract values explicitly labeled as 'Invoice Number' or 'Invoice #'. Do NOT extract Ticket No, Reference, Order ID, or Statement numbers.")
    invoice_number_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    date: Optional[str] = Field(None, description="Date the invoice was issued")
    currency: Optional[str] = Field(None, description="3-letter currency code")
    
    subtotal: Optional[float] = Field(None, description="The subtotal amount before taxes and shipping are added.")
    
    taxes: List[TaxItem] = Field(default_factory=list, description="STRICT RULE: Extract individual taxes ONLY from the summary block at the bottom of the invoice. CRITICAL GUARDRAIL: Before adding a tax here, verify you did not already extract it in `line_items`. No duplicates allowed. If a tax is already a line item, skip it here.")
    
    additional_fees: List[FeeItem] = Field(default_factory=list, description="STRICT RULE: ONLY extract fees from the summary block at the bottom. CRITICAL GUARDRAIL: Before adding a fee here, verify you did not already extract it in `line_items`. No duplicates allowed.")
    
    shipping_name: Optional[str] = Field(None, description="The exact printed name of the shipping charge (e.g., 'Freight', 'Handling', 'Delivery Fee').")
    shipping_handling: Optional[float] = Field(0.0, description="STRICT RULE: ONLY extract this if it appears in the final summary block at the bottom of the invoice, AFTER the main line items and subtotal.")
    
    total_amount: float = Field(description="Final total amount charged on the invoice")
    total_amount_confidence: Optional[str] = Field(None, description="'High', 'Medium', or 'Low'")
    
    custom_fields: Dict[str, Optional[str]] = Field(default_factory=dict, description="Extract any custom fields requested by the user. The keys MUST match exactly.")
    
    line_items: List[LineItem] = Field(description="List of all individual items purchased. CRITICAL GUARDRAIL: You must extract EVERY SINGLE ROW from the invoice table. Do NOT skip, summarize, or abbreviate rows. You must capture 100% of the items to ensure accounting math is perfectly accurate.")

class InvoiceDocument(BaseModel):
    invoices: List[InvoiceData] = Field(description="List of distinct invoices. STRICT PAGING RULE: 1) If an invoice table extends across multiple pages but shares the SAME Invoice Number, MERGE all line items into ONE invoice record. 2) If the Invoice Number CHANGES on a new page, split it into a NEW separate invoice record in this list.")

# ==============================================================
# 2. PDF to Image Conversion & Extraction
# ==============================================================

def pdf_to_base64_images(file_bytes: bytes, max_pages: int, dpi: int) -> Tuple[List[str], int]:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    base64_images = []
    total_pages = len(doc)
    if total_pages > max_pages:
        st.warning(f"PDF is {total_pages} pages long. Truncating to first {max_pages} pages based on configuration.")
    pages_to_process = min(total_pages, max_pages)
    for page_num in range(pages_to_process):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=dpi) 
        base64_images.append(base64.b64encode(pix.tobytes("jpeg")).decode('utf-8'))
    return base64_images, total_pages

def extract_invoice_data(client, deployment: str, file_bytes: bytes, custom_fields_dict: Dict[str, str], standard_aliases_dict: Dict[str, str], max_pages: int, dpi: int) -> Tuple[InvoiceDocument, int]:
    base64_images, total_pages = pdf_to_base64_images(file_bytes, max_pages, dpi)
    
    system_prompt = "You are an expert accountant processing a document that may contain multiple distinct invoices. STRICT PAGING RULES: 1) If the same invoice number continues across multiple pages, combine all line items, taxes, and totals into a SINGLE invoice record. 2) If you see a NEW invoice number, start a NEW invoice record. CRITICAL RULE AGAINST DUPLICATES: Never extract the same tax or fee twice. ANTI-LAZINESS RULE: DO NOT BE LAZY. You must extract every single line item row by row. Skipping the middle of a table, abbreviating, or summarizing items is a CRITICAL FAILURE. Your extracted total for EACH invoice must mathematically equal the calculated sum of unique items for that invoice."
    
    if standard_aliases_dict:
        alias_str = "\n".join([f"- {k}: Also look for '{v}'" for k, v in standard_aliases_dict.items()])
        system_prompt += f"\n\nSTANDARD FIELD ALIASES:\nThe user has provided common aliases for standard schema fields. Use these to help locate the data:\n{alias_str}"
    
    if custom_fields_dict:
        rules_str = "\n".join([f"- Field Key: '{k}' | Definition & Aliases: {v}" for k, v in custom_fields_dict.items()])
        system_prompt += f"\n\nSTRICT RULE: The user has requested custom data extraction based on specific definitions and alias mappings. You MUST search the invoice for the following logic and place the results in the 'custom_fields' dictionary using the EXACT 'Field Key' provided:\n{rules_str}\nIf a field or its aliases are not found, leave its value as null."

    content_array = [{"type": "text", "text": "Extract data from this multi-page document. Pay close attention to invoice numbers. Evaluate confidence ('High', 'Medium', 'Low'). Note page numbers."}]
    for img_base64 in base64_images:
        content_array.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}})

    response = client.chat.completions.create(
        model=deployment, 
        response_model=InvoiceDocument, 
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": content_array}]
    )
    return response, total_pages

def setup_excel_workbook(custom_cols: List[str]):
    wb = openpyxl.Workbook()
    ws_details = wb.active
    ws_details.title = "Invoice Details"
    details_headers = [
        "File Name", "Page #", "Vendor Name", "Vendor Address", "Bill To", "Remit To",
        "Origin", "Destination", "Invoice Number", "Date", "Currency", 
        "Material", "Description", "Quantity", "UOM", "Unit Price", "Line Total", "Subtotal", "Invoice Total",
        "Inv# Conf", "Origin Conf", "Dest Conf", "UOM Conf", "Total Conf", "Status", "Reason for Review"
    ]
    if custom_cols: details_headers.extend(custom_cols)
    ws_details.append(details_headers)
    
    ws_qc = wb.create_sheet(title="QC Summary")
    qc_headers = [
        "File Name", "Vendor Name", "Invoice Number", "Origin", "Destination", "Status", "Reason for Review", 
        "Extracted Total", "Calculated Sum", "Variance"
    ]
    ws_qc.append(qc_headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws_details, ws_qc]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
    return wb, ws_details, ws_qc

# ==============================================================
# 3. Streamlit UI & Core Logic Orchestrator
# ==============================================================

if 'extraction_details' not in st.session_state:
    st.session_state.extraction_details = None
if 'extraction_summary' not in st.session_state:
    st.session_state.extraction_summary = None
if 'extraction_excel' not in st.session_state:
    st.session_state.extraction_excel = None
if 'start_processing' not in st.session_state:
    st.session_state.start_processing = False
if 'files_to_process' not in st.session_state:
    st.session_state.files_to_process = []

def run_extraction_process(files_list, custom_fields_dict, standard_aliases_dict, max_pages, dpi, sleep_time, prefix=""):
    client = instructor.from_openai(AzureOpenAI(azure_endpoint=AZURE_ENDPOINT, api_key=AZURE_API_KEY, api_version=AZURE_API_VERSION))
    
    custom_col_keys = list(custom_fields_dict.keys())
    wb, ws_details, ws_qc = setup_excel_workbook(custom_col_keys)
    
    # ---> NEW: Load Master CSV directly into memory for rapid mapping <---
    master_supp_df = load_master_suppliers()
    # Create O(1) lookup dictionary matching lowercase raw names to standard clean names
    supplier_map = dict(zip(master_supp_df['Raw_Vendor_Name'].astype(str).str.lower(), master_supp_df['Clean_Supplier_Name']))
    
    success_count, error_count = 0, 0
    progress_bar = st.progress(0, text=f"Initializing {prefix} processing sequence...")
    status_text = st.empty()
    
    current_run_summary = []
    current_run_details = []

    start_time_batch = time.time()
    total_files = len(files_list)
    current_batch_id = datetime.now().strftime("BATCH_%Y%m%d_%H%M%S")

    for idx, file in enumerate(files_list):
        filename = file.name
        now = datetime.now()
        current_date, current_time = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")

        status_text.markdown(f"**Extracting ({idx+1}/{total_files}):** `{filename}`...")
        file_start_time = time.time()
        
        try:
            file.seek(0)
            file_bytes = file.read()
            extracted_document, total_pages = extract_invoice_data(client, AZURE_DEPLOYMENT, file_bytes, custom_fields_dict, standard_aliases_dict, max_pages, dpi)
            file_proc_time = round(time.time() - file_start_time, 2)
            
            if not extracted_document.invoices:
                ws_qc.append([filename, "N/A", "N/A", "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found in PDF", "", "", ""])
                insert_audit_record((current_date, current_time, filename, "N/A", "N/A", "N/A", "N/A", "FAIL - NO DATA", "0 Invoices Found", 0.0, 0.0, 0.0, file_proc_time, total_pages, current_batch_id, "N/A"))
                current_run_summary.append({"File Name": filename, "Vendor Name": "N/A", "Invoice #": "N/A", "Origin": "N/A", "Destination": "N/A", "Status": "FAIL", "Reason": "No Invoices Found in PDF"})
                continue
            
            for extracted_data in extracted_document.invoices:
                # ---> NEW: Dynamically cross-reference Master CSV <---
                raw_vendor = extracted_data.vendor_name
                clean_supp = supplier_map.get(str(raw_vendor).lower())
                if not clean_supp:
                    clean_supp = standardize_vendor(raw_vendor)
                    
                calculated_line_sum = sum(item.line_total for item in extracted_data.line_items if item.line_total is not None)
                safe_ship = extracted_data.shipping_handling or 0.0
                safe_total_tax = sum(tax.tax_amount for tax in extracted_data.taxes if tax.tax_amount is not None)
                safe_fees = sum(fee.fee_amount for fee in extracted_data.additional_fees if fee.fee_amount is not None)
                
                total_calculated = calculated_line_sum + safe_total_tax + safe_ship + safe_fees
                variance = round(extracted_data.total_amount - total_calculated, 2)
                
                review_reasons = []
                if variance != 0.0: review_reasons.append(f"Math Variance of {variance}")
                if extracted_data.invoice_number_confidence == "Low": review_reasons.append("Low Conf: Invoice #")
                if extracted_data.origin_confidence == "Low": review_reasons.append("Low Conf: Origin")
                if extracted_data.destination_confidence == "Low": review_reasons.append("Low Conf: Destination")
                if extracted_data.total_amount_confidence == "Low": review_reasons.append("Low Conf: Total Amount")
                if len(extracted_data.line_items) == 0: review_reasons.append("Missing: 0 Line Items Found")
                
                for item in extracted_data.line_items:
                    if item.uom_confidence == "Low":
                        short_desc = item.description[:15] + "..." if item.description and len(item.description) > 15 else item.description
                        review_reasons.append(f"Low Conf: UOM on '{short_desc}'")
                
                needs_review = len(review_reasons) > 0
                status = "FAIL - NEEDS REVIEW" if needs_review else "PASS"
                reasons_string = " | ".join(review_reasons) if needs_review else "N/A"
                
                def create_row_dict(page_num, material, desc, qty, uom, uom_conf, price, line_total, line_orig=None, line_dest=None):
                    final_origin = line_orig if line_orig else extracted_data.origin
                    final_dest = line_dest if line_dest else extracted_data.destination
                    
                    row_data = {
                        "File Name": filename, "Page #": page_num,
                        "Vendor Name": extracted_data.vendor_name, "Vendor Address": extracted_data.vendor_address,
                        "Bill To": extracted_data.bill_to, "Remit To": extracted_data.remit_to,
                        "Origin": final_origin, "Destination": final_dest, 
                        "Invoice Number": extracted_data.invoice_number, "Date": extracted_data.date, "Currency": extracted_data.currency,
                        "Material": material, "Description": desc, 
                        "Quantity": qty, "UOM": uom, "Unit Price": price, "Line Total": line_total,
                        "Subtotal": extracted_data.subtotal, "Invoice Total": extracted_data.total_amount,
                        "Inv# Conf": extracted_data.invoice_number_confidence, "Origin Conf": extracted_data.origin_confidence,
                        "Dest Conf": extracted_data.destination_confidence, "UOM Conf": uom_conf, "Total Conf": extracted_data.total_amount_confidence,
                        "Status": status, "Reason for Review": reasons_string
                    }
                    for col in custom_col_keys:
                        row_data[col] = extracted_data.custom_fields.get(col, "Not Found")
                    return row_data

                if len(extracted_data.line_items) == 0:
                    row_dict = create_row_dict(None, None, "NO ITEMS FOUND", None, None, None, None, 0.0)
                    ws_details.append(list(row_dict.values()))
                    current_run_details.append(row_dict)
                else:
                    for item in extracted_data.line_items:
                        row_dict = create_row_dict(
                            item.page_number, item.material, item.description, item.quantity, 
                            item.uom, item.uom_confidence, item.unit_price, item.line_total, 
                            line_orig=item.line_origin, line_dest=item.line_destination
                        )
                        ws_details.append(list(row_dict.values()))
                        current_run_details.append(row_dict)
                        
                if safe_ship > 0: 
                    row_dict = create_row_dict(None, "SHIPPING", extracted_data.shipping_name or "Shipping", None, None, None, None, safe_ship)
                    ws_details.append(list(row_dict.values()))
                    current_run_details.append(row_dict)
                    
                for tax in extracted_data.taxes:
                    if tax.tax_amount is not None and tax.tax_amount > 0: 
                        row_dict = create_row_dict(None, "TAX", tax.tax_name, None, None, None, None, tax.tax_amount)
                        ws_details.append(list(row_dict.values()))
                        current_run_details.append(row_dict)
                        
                for fee in extracted_data.additional_fees:
                    if fee.fee_amount is not None and fee.fee_amount > 0: 
                        row_dict = create_row_dict(None, "FEE", fee.fee_name, None, None, None, None, fee.fee_amount)
                        ws_details.append(list(row_dict.values()))
                        current_run_details.append(row_dict)
                
                first_line_orig = extracted_data.line_items[0].line_origin if len(extracted_data.line_items) > 0 else None
                first_line_dest = extracted_data.line_items[0].line_destination if len(extracted_data.line_items) > 0 else None
                qc_origin = first_line_orig if first_line_orig else extracted_data.origin
                qc_dest = first_line_dest if first_line_dest else extracted_data.destination

                ws_qc.append([
                    filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                    qc_origin, qc_dest,
                    status, reasons_string, extracted_data.total_amount, total_calculated, variance
                ])
                
                # ---> NEW: Database now correctly saves the auto-mapped clean_supp <---
                insert_audit_record((
                    current_date, current_time, filename, extracted_data.vendor_name, extracted_data.invoice_number, 
                    str(qc_origin), str(qc_dest), status, reasons_string, 
                    extracted_data.total_amount, total_calculated, variance, file_proc_time, total_pages, current_batch_id, clean_supp
                ))

                current_run_summary.append({
                    "File Name": filename, "Vendor Name": extracted_data.vendor_name, "Invoice #": extracted_data.invoice_number,
                    "Origin": qc_origin or "Missing", "Destination": qc_dest or "Missing",
                    "Variance": f"${variance:,.2f}", "Status": "✅ PASS" if status == "PASS" else "⚠️ FAIL",
                    "Proc Time": f"{file_proc_time}s"
                })
                success_count += 1
                
        except Exception as e:
            file_proc_time = round(time.time() - file_start_time, 2)
            insert_audit_record((current_date, current_time, filename, "N/A", "N/A", "N/A", "N/A", "FAIL - CRITICAL ERROR", str(e), 0.0, 0.0, 0.0, file_proc_time, 0, current_batch_id, "ERROR"))
            current_run_summary.append({"File Name": filename, "Vendor Name": "ERROR", "Invoice #": "ERROR", "Origin": "ERROR", "Destination": "ERROR", "Status": "❌ ERROR", "Reason": "API/System Crash", "Proc Time": f"{file_proc_time}s"})
            error_count += 1
        
        if idx < total_files - 1: time.sleep(sleep_time) 
        
        files_processed = idx + 1
        elapsed_time = time.time() - start_time_batch
        avg_time_per_file = elapsed_time / files_processed
        remaining_files = total_files - files_processed
        mins, secs = divmod(int(avg_time_per_file * remaining_files), 60)
        eta_string = f" | ETA: {mins}m {secs}s" if (avg_time_per_file * remaining_files) > 0 else ""
        progress_bar.progress(files_processed / total_files, text=f"Processing {files_processed}/{total_files} documents{eta_string}")

    fetch_audit_data.clear()

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    st.session_state.extraction_details = current_run_details
    st.session_state.extraction_summary = current_run_summary
    st.session_state.extraction_excel = excel_buffer
    
    progress_bar.empty()
    status_text.empty()
    st.success(f"🎉 {prefix} Batch Processing Complete! Invoices Extracted: {success_count} | Errors: {error_count}")


# --- DB & UI Dialogs ---
@st.dialog("⚠️ Duplicate Files Detected")
def confirm_duplicates_dialog(duplicate_files, unique_files):
    st.warning(f"Found {len(duplicate_files)} file(s) that have already been processed and exist in the master database.")
    for f in duplicate_files:
        st.write(f"- `{f.name}`")
    st.write("Do you want to extract these again? (This will add duplicate entries to your audit logs).")
    
    col1, col2 = st.columns(2)
    if col1.button("Yes, Process All", type="primary"):
        st.session_state.files_to_process = unique_files + duplicate_files
        st.session_state.start_processing = True
        st.rerun()
    if col2.button("No, Unique Only"):
        if unique_files:
            st.session_state.files_to_process = unique_files
            st.session_state.start_processing = True
            st.rerun()
        else:
            st.warning("No unique files to process.")
            time.sleep(2)
            st.rerun()

@st.dialog("🗑️ Remove Duplicate Invoices")
def dialog_remove_duplicates():
    st.warning("This will permanently delete all duplicate records (keeping the oldest one) based on Vendor Name + Invoice Number.")
    user_name = st.text_input("Enter your name to authorize deletion:")
    if st.button("Confirm Deletion", type="primary"):
        if not user_name.strip():
            st.error("Name is required for the audit log.")
        else:
            remove_duplicates_db(user_name.strip())
            st.rerun()

@st.dialog("⚠️ Wipe Master Database")
def dialog_wipe_db():
    st.error("CRITICAL WARNING: This will permanently delete ALL extracted invoice data from the system.")
    user_name = st.text_input("Enter your name to authorize complete wipe:")
    if st.button("Confirm Complete Wipe", type="primary"):
        if not user_name.strip():
            st.error("Name is required for the audit log.")
        else:
            wipe_master_db(user_name.strip())
            st.rerun()

@st.dialog("📄 Fullscreen Document Viewer", width="large")
def view_fullscreen_pdf(file_bytes, file_name):
    st.markdown(f"### {file_name}")
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    
    html_content = '<div style="height: 75vh; overflow-y: scroll; border: 1px solid #ddd; padding: 20px; background-color: #525659; text-align: center;">'
    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(dpi=150)
        img_b64 = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
        html_content += f'<img src="data:image/jpeg;base64,{img_b64}" style="width: 95%; max-width: 900px; margin-bottom: 20px; box-shadow: 0px 4px 10px rgba(0,0,0,0.5);"><br>'
    html_content += '</div>'
    
    st.markdown(html_content, unsafe_allow_html=True)


# ==============================================================
# 4. Streamlit App Layout
# ==============================================================

st.set_page_config(page_title="AI Invoice Intelligence", page_icon="🧾", layout="wide")
st.title("🧾 AI Invoice Intelligence Platform")

# Sidebar
with st.sidebar:
    st.header("📄 File Upload")
    uploaded_files = st.file_uploader("Upload PDF Documents", type=["pdf"], accept_multiple_files=True)
    
    st.divider()
    
    with st.expander("⚙️ Processing Configuration", expanded=False):
        config_max_pages = st.number_input("Max Pages per PDF", min_value=1, max_value=100, value=15)
        config_dpi = st.slider("Render Resolution (DPI)", min_value=72, max_value=600, value=300, step=72)
        config_sleep_time = st.number_input("API Sleep Time (seconds)", min_value=0, max_value=60, value=3)

    with st.expander("🎯 Standard Field Aliases", expanded=False):
        st.write("Add alternative names (comma-separated) for built-in fields to help the AI find them.")
        default_standard = pd.DataFrame({
            "Standard Field": [
                "invoice_number", "date", "vendor_name", "vendor_address", 
                "bill_to", "remit_to", "origin", "destination", "currency", 
                "subtotal", "shipping_name", "shipping_handling", "total_amount",
                "material", "description", "line_origin", "line_destination", "quantity", "uom", "unit_price", "line_total",
                "tax_name", "tax_amount", "fee_name", "fee_amount"
            ],
            "Aliases": [""] * 25
        })
        standard_df = st.data_editor(default_standard, disabled=["Standard Field"], use_container_width=True, hide_index=True)
        
        standard_aliases_dict = {}
        for _, row in standard_df.iterrows():
            if str(row["Aliases"]).strip() and str(row["Aliases"]) != "None":
                standard_aliases_dict[row["Standard Field"]] = str(row["Aliases"]).strip()
                
    with st.expander("🎯 New Custom Fields", expanded=False):
        st.write("Add entirely new columns to extract.")
        default_custom = pd.DataFrame({
            "Field Name": [""],
            "Description & Aliases": [""]
        })
        custom_df = st.data_editor(default_custom, num_rows="dynamic", use_container_width=True, hide_index=True)
        
        custom_fields_dict = {}
        for _, row in custom_df.iterrows():
            name = str(row["Field Name"]).strip()
            desc = str(row["Description & Aliases"]).strip()
            if name and name != "None":
                custom_fields_dict[name] = desc

# Tabs
tab_extract, tab_viewer, tab_batch, tab_analytics, tab_system = st.tabs([
    "⚙️ Extraction Suite", "📄 Document Viewer", "✅ Current Batch QA", "📊 Business Analytics", "🤖 System & Drift Analysis"
])

# ---------------------------------------------------------
# TAB 1: EXTRACTION SUITE
# ---------------------------------------------------------
with tab_extract:
    st.write("Upload invoices via the sidebar and click below to process them. Results will appear here instantly.")
    
    if st.button("🚀 Start Extraction", type="primary"):
        if not uploaded_files:
            st.error("Please upload at least one PDF file from the sidebar.")
        elif not all([AZURE_ENDPOINT, AZURE_API_KEY, AZURE_DEPLOYMENT]):
            st.error("⚠️ Azure credentials missing. Check your `.env` file.")
        else:
            df_audit_existing = fetch_audit_data()
            existing_filenames = df_audit_existing['file_name'].unique().tolist() if not df_audit_existing.empty else []
            
            unique_files = [f for f in uploaded_files if f.name not in existing_filenames]
            duplicate_files = [f for f in uploaded_files if f.name in existing_filenames]
            
            if duplicate_files:
                confirm_duplicates_dialog(duplicate_files, unique_files)
            else:
                st.session_state.files_to_process = unique_files
                st.session_state.start_processing = True

    if st.session_state.start_processing:
        run_extraction_process(
            st.session_state.files_to_process, 
            custom_fields_dict, 
            standard_aliases_dict, 
            config_max_pages, 
            config_dpi,
            config_sleep_time
        )
        st.session_state.start_processing = False
        st.rerun()

    if st.session_state.extraction_details is not None:
        st.subheader("📝 Line Item Details (Full Extraction)")
        st.dataframe(pd.DataFrame(st.session_state.extraction_details), use_container_width=True, hide_index=True)

        st.divider()

        st.subheader("🛡️ QC Summary")
        st.dataframe(pd.DataFrame(st.session_state.extraction_summary), use_container_width=True, hide_index=True)

        st.download_button(
            label="📥 Download Excel Report",
            data=st.session_state.extraction_excel,
            file_name=f"Extraction_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ---------------------------------------------------------
# TAB 2: DOCUMENT VIEWER
# ---------------------------------------------------------
with tab_viewer:
    st.header("📄 Document Viewer")
    
    if uploaded_files:
        st.write("Search and click '🔍 View Full Screen' to examine a document in detail.")
        
        search_query = st.text_input("🔍 Search by File Name", "").lower()
        filtered_files = [f for f in uploaded_files if search_query in f.name.lower()]
        
        if not filtered_files:
            st.warning("No files match your search query.")
        else:
            cols = st.columns(4)
            for idx, file in enumerate(filtered_files):
                col = cols[idx % 4]
                with col:
                    st.markdown(f"<div style='white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-size: 14px; margin-bottom: 5px;' title='{file.name}'><b>{file.name}</b></div>", unsafe_allow_html=True)
                    
                    file.seek(0)
                    doc = fitz.open(stream=file.read(), filetype="pdf")
                    if len(doc) > 0:
                        page = doc[0]
                        pix = page.get_pixmap(dpi=72)
                        img_b64 = base64.b64encode(pix.tobytes("jpeg")).decode('utf-8')
                        st.markdown(f'<img src="data:image/jpeg;base64,{img_b64}" style="width: 100%; border: 1px solid #ccc; box-shadow: 2px 2px 5px rgba(0,0,0,0.1); margin-bottom: 10px;">', unsafe_allow_html=True)
                    
                    file.seek(0)
                    if st.button("🔍 View Full Screen", key=f"view_{idx}", use_container_width=True):
                        view_fullscreen_pdf(file.read(), file.name)
    else:
        st.info("Upload PDF documents in the sidebar to view them here.")

# ---------------------------------------------------------
# TAB 3: CURRENT BATCH QA & SUPPLIER MAPPING
# ---------------------------------------------------------
with tab_batch:
    st.header("✅ Current Batch QA & Master Data Preparation")
    st.write("Review the quality of the most recent extraction batch, map messy vendor names to Clean Suppliers, and execute AI Routing Suggestions.")
    
    df_audit = fetch_audit_data()
    
    if df_audit.empty:
        st.info("Process a batch of invoices to view current metrics.")
    else:
        if 'batch_id' in df_audit.columns and not df_audit['batch_id'].isna().all():
            latest_batch = df_audit['batch_id'].dropna().iloc[-1]
            batch_df = df_audit[df_audit['batch_id'] == latest_batch].copy()
            st.markdown(f"**Latest Batch ID:** `{latest_batch}` | **Total Invoices in Batch:** `{len(batch_df)}`")
        else:
            latest_date = df_audit['extraction_date'].max()
            batch_df = df_audit[df_audit['extraction_date'] == latest_date].copy()
            st.markdown(f"**Latest Batch Date:** `{latest_date}` | **Total Invoices in Batch:** `{len(batch_df)}`")
        
        missing_flags = ['N/A', 'None', '', 'null', 'None']
        batch_routing_issues = batch_df[
            batch_df['origin'].isin(missing_flags) | batch_df['origin'].isna() |
            batch_df['destination'].isin(missing_flags) | batch_df['destination'].isna()
        ]
        batch_mismatches = batch_df[batch_df['variance'] != 0.0]
        
        bc1, bc2, bc3, bc4 = st.columns(4)
        bc1.metric("Batch Processing Volume", f"{len(batch_df)}")
        bc2.metric("Batch Math Mismatches", f"{len(batch_mismatches)}", delta_color="inverse")
        bc3.metric("Batch Routing Issues", f"{len(batch_routing_issues)}", delta_color="inverse")
        batch_acc = ((len(batch_df) - len(batch_mismatches)) / len(batch_df)) * 100 if len(batch_df) > 0 else 0
        bc4.metric("Batch Math Accuracy", f"{batch_acc:.1f}%")

        st.divider()

        st.subheader("1. Clean Supplier Mapping")
        st.write("Map raw extracted vendor names to standard 'Clean Suppliers'. Saving updates both the Database and the Master CSV.")
        
        if 'clean_supplier' in batch_df.columns:
            mapping_df = batch_df[['id', 'file_name', 'vendor_name', 'clean_supplier']].copy()
            
            edited_mapping = st.data_editor(
                mapping_df, 
                use_container_width=True, 
                hide_index=True,
                column_config={
                    "id": st.column_config.NumberColumn("DB ID", disabled=True),
                    "file_name": st.column_config.TextColumn("File Name", disabled=True),
                    "vendor_name": st.column_config.TextColumn("Extracted Raw Vendor", disabled=True),
                    "clean_supplier": st.column_config.TextColumn("✏️ Edit Clean Supplier Name")
                }
            )
            
            if st.button("💾 Save Mappings & Update Master CSV", type="primary"):
                master_df = load_master_suppliers()
                new_mappings = []
                
                for index, row in edited_mapping.iterrows():
                    update_audit_record_supplier(row['id'], row['clean_supplier'])
                    new_mappings.append({"Raw_Vendor_Name": row['vendor_name'], "Clean_Supplier_Name": row['clean_supplier']})
                
                # Merge new mappings with existing master CSV
                new_df = pd.DataFrame(new_mappings)
                combined_df = pd.concat([master_df, new_df]).drop_duplicates(subset=['Raw_Vendor_Name'], keep='last')
                save_master_suppliers(combined_df)
                
                st.success("Successfully updated supplier mappings and Master CSV!")
                st.rerun()

            st.divider()
            
            st.subheader("2. AI Routing Suggestions (HitL Framework)")
            st.write("Scan the Master Database to suggest the Top 5 most frequent historical routes for invoices missing data.")
            
            if st.button("⚙️ Generate Top 5 Historical Suggestions"):
                valid_origins = df_audit[~df_audit['origin'].isin(missing_flags) & df_audit['origin'].notna()]
                valid_dests = df_audit[~df_audit['destination'].isin(missing_flags) & df_audit['destination'].notna()]
                
                orig_counts = valid_origins.groupby(['clean_supplier', 'origin']).size().reset_index(name='count')
                orig_counts = orig_counts.sort_values(['clean_supplier', 'count'], ascending=[True, False])
                top_origins = orig_counts.groupby('clean_supplier').head(5)
                
                orig_dict = {}
                for supp, group in top_origins.groupby('clean_supplier'):
                    suggs = [f"{idx+1}. {row['origin']} (x{row['count']})" for idx, row in enumerate(group.to_dict('records'))]
                    orig_dict[supp] = " | ".join(suggs)
                    
                dest_counts = valid_dests.groupby(['clean_supplier', 'destination']).size().reset_index(name='count')
                dest_counts = dest_counts.sort_values(['clean_supplier', 'count'], ascending=[True, False])
                top_dests = dest_counts.groupby('clean_supplier').head(5)
                
                dest_dict = {}
                for supp, group in top_dests.groupby('clean_supplier'):
                    suggs = [f"{idx+1}. {row['destination']} (x{row['count']})" for idx, row in enumerate(group.to_dict('records'))]
                    dest_dict[supp] = " | ".join(suggs)
                
                for idx, row in batch_routing_issues.iterrows():
                    supp = row['clean_supplier'] if pd.notna(row['clean_supplier']) else standardize_vendor(row['vendor_name'])
                    sug_orig = orig_dict.get(supp, "No historical data")
                    sug_dest = dest_dict.get(supp, "No historical data")
                    update_audit_suggestions(row['id'], sug_orig, sug_dest)
                    
                st.success("Historical suggestions generated successfully!")
                st.rerun()
                
            st.write("Review the Top 5 historical suggestions below and type your final choice into the Approved column.")
            
            current_db = fetch_audit_data()
            if 'batch_id' in current_db.columns and not current_db['batch_id'].isna().all():
                c_latest_batch = current_db['batch_id'].dropna().iloc[-1]
                current_batch_issues = current_db[(current_db['batch_id'] == c_latest_batch) & (
                    current_db['origin'].isin(missing_flags) | current_db['origin'].isna() |
                    current_db['destination'].isin(missing_flags) | current_db['destination'].isna()
                )].copy()
            else:
                current_batch_issues = current_db[(current_db['extraction_date'] == latest_date) & (
                    current_db['origin'].isin(missing_flags) | current_db['origin'].isna() |
                    current_db['destination'].isin(missing_flags) | current_db['destination'].isna()
                )].copy()

            if not current_batch_issues.empty:
                review_df = current_batch_issues[['id', 'file_name', 'clean_supplier', 'origin', 'suggested_origin', 'final_origin', 'destination', 'suggested_destination', 'final_destination']]
                
                edited_routes = st.data_editor(
                    review_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "id": st.column_config.NumberColumn("DB ID", disabled=True),
                        "file_name": st.column_config.TextColumn("File", disabled=True),
                        "clean_supplier": st.column_config.TextColumn("Supplier", disabled=True),
                        "origin": st.column_config.TextColumn("Raw Origin", disabled=True),
                        "suggested_origin": st.column_config.TextColumn("🤖 Suggested Origins (Top 5)", disabled=True),
                        "final_origin": st.column_config.TextColumn("✅ Approved Origin"),
                        "destination": st.column_config.TextColumn("Raw Dest", disabled=True),
                        "suggested_destination": st.column_config.TextColumn("🤖 Suggested Dests (Top 5)", disabled=True),
                        "final_destination": st.column_config.TextColumn("✅ Approved Dest")
                    }
                )
                
                if st.button("💾 Save Approved Routes to DB", type="primary"):
                    for index, row in edited_routes.iterrows():
                        if pd.notna(row['final_origin']) or pd.notna(row['final_destination']):
                            f_orig = row['final_origin'] if pd.notna(row['final_origin']) else None
                            f_dest = row['final_destination'] if pd.notna(row['final_destination']) else None
                            update_audit_final_routes(row['id'], f_orig, f_dest)
                    st.success("Final routes locked into the Master Database.")
                    st.rerun()
            else:
                st.info("No routing issues found in the current batch!")

            st.divider()
            
            # ---> NEW: View & Edit Master CSV Directly <---
            st.subheader("📂 Master Supplier Database (CSV)")
            st.write("View or manually override the global mapping rules. These apply automatically to future extractions.")
            master_df = load_master_suppliers()
            edited_master = st.data_editor(master_df, num_rows="dynamic", use_container_width=True)
            if st.button("💾 Update Master CSV Directly"):
                save_master_suppliers(edited_master)
                st.success("Master CSV Rules Updated!")

# ---------------------------------------------------------
# TAB 4: BUSINESS ANALYTICS
# ---------------------------------------------------------
with tab_analytics:
    st.header("📊 Business Analytics - Extraction Quality")
    df_audit = fetch_audit_data()
    
    if df_audit.empty:
        st.info("📭 No data available. Process some invoices in the Extraction Suite to generate analytics.")
    else:
        # We use clean_supplier for analytics to ensure consistency
        if 'clean_supplier' in df_audit.columns:
            df_audit['vendor_name'] = df_audit['clean_supplier'].fillna(df_audit['vendor_name'].apply(standardize_vendor))
        else:
            df_audit['vendor_name'] = df_audit['vendor_name'].apply(standardize_vendor)
            
        missing_flags = ['N/A', 'None', '', 'null', 'None']
        
        valid_df = df_audit[~df_audit['invoice_number'].isin(['N/A', 'ERROR', '', None]) & ~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]
        duplicates_df = valid_df[valid_df.duplicated(subset=['vendor_name', 'invoice_number'], keep='first')]
        amount_mismatches_df = df_audit[df_audit['variance'] != 0.0]
        routing_issues_df = df_audit[df_audit['origin'].isin(missing_flags) | df_audit['origin'].isna() | df_audit['destination'].isin(missing_flags) | df_audit['destination'].isna()]
        missing_origin_df = df_audit[df_audit['origin'].isin(missing_flags) | df_audit['origin'].isna()]
        missing_dest_df = df_audit[df_audit['destination'].isin(missing_flags) | df_audit['destination'].isna()]

        # 1. KPIs
        st.subheader("Invoice & Vendor Overview")
        col1, col2, col3, col4, col5 = st.columns(5)
        
        total_invoices = len(df_audit)
        total_vendors = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]['vendor_name'].nunique()
        total_spend = df_audit['extracted_total'].sum()
        dup_count = len(duplicates_df)
        dup_spend = duplicates_df['extracted_total'].sum()

        col1.metric("Total Invoices", f"{total_invoices:,}")
        col2.metric("Total Unique Vendors", f"{total_vendors:,}")
        col3.metric("Total Spend", f"${total_spend / 1_000_000:,.2f}M")
        col4.metric("⚠️ Duplicates Found", f"{dup_count:,}", delta_color="inverse")
        col5.metric("⚠️ Duplicate Value", f"${dup_spend / 1_000_000:,.2f}M", delta_color="inverse")

        st.divider()

        # 2. Quality Scatter & Box Plot
        st.subheader("🎯 Extraction Quality & Outlier Detection")
        st.write("Identifies severe mathematical anomalies where the LLM's extracted total deviates wildly from the raw line items.")
        
        c1, c2 = st.columns(2)
        with c1:
            fig_scatter = px.scatter(
                df_audit[df_audit['status'] != 'FAIL - CRITICAL ERROR'], 
                x='calculated_sum', y='extracted_total', 
                color='vendor_name', hover_data=['file_name', 'variance'],
                title="Math Validity: Expected vs. Extracted Total"
            )
            max_val = max(df_audit['calculated_sum'].max(), df_audit['extracted_total'].max())
            fig_scatter.add_trace(go.Scatter(x=[0, max_val], y=[0, max_val], mode='lines', line=dict(color='black', dash='dash'), name='Perfect Match'))
            fig_scatter.update_layout(xaxis_title="Calculated Sum (Items + Tax + Fees)", yaxis_title="Extracted Total")
            st.plotly_chart(fig_scatter, use_container_width=True)
            
        with c2:
            if not amount_mismatches_df.empty:
                fig_box = px.box(
                    amount_mismatches_df, x='vendor_name', y='variance', 
                    title="Variance Range by Vendor (Identifies erratic extraction)",
                    color_discrete_sequence=['#e74c3c']
                )
                fig_box.update_layout(xaxis_title="Vendor", yaxis_title="Variance Magnitude ($)")
                st.plotly_chart(fig_box, use_container_width=True)
            else:
                st.success("No math variance detected across vendors.")

        st.divider()
        
        # 3. Document Complexity Impact
        st.subheader("📄 Document Complexity Impact")
        st.write("Analyzes if longer documents lead to a higher failure rate in data extraction.")
        
        if 'page_count' in df_audit.columns:
            df_audit['Clean_Status'] = df_audit['status'].apply(lambda x: 'PASS' if x == 'PASS' else 'FAIL')
            complexity_df = df_audit.groupby(['page_count', 'Clean_Status']).size().reset_index(name='count')
            
            fig_complex = px.bar(
                complexity_df, x='page_count', y='count', color='Clean_Status', 
                barmode='stack', color_discrete_map={'PASS': '#2ecc71', 'FAIL': '#e74c3c'}
            )
            fig_complex.update_layout(xaxis_title="Number of Pages in Document", yaxis_title="Number of Invoices")
            st.plotly_chart(fig_complex, use_container_width=True)
        else:
            st.info("Page count analytics will populate on new extractions.")
            
        st.divider()

        # 4. Filterable Vendor Routing & Exception Breakdown
        st.subheader("🔎 Vendor Routing & Exception Breakdown")
        st.write("Filter to see a detailed routing scorecard and math error count by vendor.")
        
        vendor_df = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].copy()
        
        vendor_df['missing_origin'] = vendor_df['origin'].isin(missing_flags) | vendor_df['origin'].isna()
        vendor_df['valid_origin'] = ~vendor_df['missing_origin']
        vendor_df['missing_dest'] = vendor_df['destination'].isin(missing_flags) | vendor_df['destination'].isna()
        vendor_df['valid_dest'] = ~vendor_df['missing_dest']
        vendor_df['amount_mismatch'] = vendor_df['variance'] != 0.0

        vendor_summary = vendor_df.groupby('vendor_name').agg(
            Total_Invoices=('id', 'count'),
            Valid_Origin=('valid_origin', 'sum'),
            Missing_Origin=('missing_origin', 'sum'),
            Valid_Destination=('valid_dest', 'sum'),
            Missing_Destination=('missing_dest', 'sum'),
            Amount_Mismatch=('amount_mismatch', 'sum')
        ).reset_index()
        
        vendor_summary.columns = [
            'Vendor Name', 'Total Invoices', 
            'Has Origin', 'Missing Origin', 
            'Has Destination', 'Missing Destination', 
            'Amount Mismatch'
        ]
        vendor_summary = vendor_summary.sort_values('Total Invoices', ascending=False)
        
        all_vendors = vendor_summary['Vendor Name'].tolist()
        selected_vendors = st.multiselect("Filter by Vendor(s):", options=all_vendors, default=[])
        
        if selected_vendors:
            filtered_summary = vendor_summary[vendor_summary['Vendor Name'].isin(selected_vendors)]
        else:
            filtered_summary = vendor_summary
            
        st.dataframe(filtered_summary, use_container_width=True, hide_index=True)

        st.divider()

        # 5. Top Vendor Charts
        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Top Vendors by Invoice Volume")
            vendor_counts = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]['vendor_name'].value_counts().reset_index().head(10)
            vendor_counts.columns = ['Vendor', 'Invoice Count']
            fig_bar = px.bar(vendor_counts, x='Invoice Count', y='Vendor', orientation='h', text_auto=True, color_discrete_sequence=['#2ecc71'])
            fig_bar.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_bar, use_container_width=True)

        with c2:
            st.subheader("Top Vendors by Financial Value ($M)")
            vendor_value = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('vendor_name')['extracted_total'].sum().reset_index()
            vendor_value['extracted_total_M'] = vendor_value['extracted_total'] / 1_000_000
            vendor_value = vendor_value.sort_values(by='extracted_total_M', ascending=False).head(10)
            vendor_value.columns = ['Vendor', 'Total Value ($)', 'Total Value (Millions)']
            fig_val = px.bar(vendor_value, x='Total Value (Millions)', y='Vendor', orientation='h', text_auto='.2f', color_discrete_sequence=['#9b59b6'])
            fig_val.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_val, use_container_width=True)
            
        st.divider()
        
        # 6. Exception Bar Charts
        st.subheader("⚠️ Vendor Exceptions & Missing Data Analysis")
        
        err1, err2, err3 = st.columns(3)
        with err1:
            st.markdown("**Vendors Missing Origin**")
            if not missing_origin_df.empty:
                mo_vendor = missing_origin_df['vendor_name'].value_counts().reset_index().head(10)
                mo_vendor.columns = ['Vendor', 'Count']
                fig_mo = px.bar(mo_vendor, x='Count', y='Vendor', orientation='h', text_auto=True, color_discrete_sequence=['#e74c3c'])
                fig_mo.update_layout(yaxis={'categoryorder':'total ascending'}, height=300, margin=dict(l=0, r=0, t=30, b=0))
                st.plotly_chart(fig_mo, use_container_width=True)
            else:
                st.success("No missing origins.")

        with err2:
            st.markdown("**Vendors Missing Destination**")
            if not missing_dest_df.empty:
                md_vendor = missing_dest_df['vendor_name'].value_counts().reset_index().head(10)
                md_vendor.columns = ['Vendor', 'Count']
                fig_md = px.bar(md_vendor, x='Count', y='Vendor', orientation='h', text_auto=True, color_discrete_sequence=['#e67e22'])
                fig_md.update_layout(yaxis={'categoryorder':'total ascending'}, height=300, margin=dict(l=0, r=0, t=30, b=0))
                st.plotly_chart(fig_md, use_container_width=True)
            else:
                st.success("No missing destinations.")

        with err3:
            st.markdown("**Vendors with Amount Mismatches**")
            if not amount_mismatches_df.empty:
                am_vendor = amount_mismatches_df['vendor_name'].value_counts().reset_index().head(10)
                am_vendor.columns = ['Vendor', 'Count']
                fig_am = px.bar(am_vendor, x='Count', y='Vendor', orientation='h', text_auto=True, color_discrete_sequence=['#f1c40f'])
                fig_am.update_layout(yaxis={'categoryorder':'total ascending'}, height=300, margin=dict(l=0, r=0, t=30, b=0))
                st.plotly_chart(fig_am, use_container_width=True)
            else:
                st.success("No math errors detected.")

        st.divider()

        # 7. Top Corridors
        st.subheader("📍 Top Shipping Corridors")
        df_valid_routes = df_audit[~df_audit['origin'].isin(missing_flags) & ~df_audit['destination'].isin(missing_flags)].copy()
        
        if not df_valid_routes.empty:
            df_valid_routes['Corridor'] = df_valid_routes['origin'].astype(str) + " ➡️ " + df_valid_routes['destination'].astype(str)
            corridor_counts = df_valid_routes['Corridor'].value_counts().reset_index().head(10)
            corridor_counts.columns = ['Shipping Corridor', 'Volume']
            fig_corridor = px.bar(corridor_counts, x='Volume', y='Shipping Corridor', orientation='h', text_auto=True, color_discrete_sequence=['#34495e'])
            fig_corridor.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_corridor, use_container_width=True)
        else:
            st.info("Not enough valid Origin & Destination pairs to map corridors.")

        st.divider()

        # 8. Hidden Expanders
        st.subheader("📋 Raw Data & Exception Tables")
        
        with st.expander("🚨 Routing Exception Report (Missing Data)", expanded=False):
            st.write("Invoices strictly missing Origin or Destination data:")
            if not routing_issues_df.empty:
                review_cols = ['file_name', 'vendor_name', 'invoice_number', 'origin', 'destination', 'reason_for_review']
                st.dataframe(routing_issues_df[review_cols], use_container_width=True, hide_index=True)
            else:
                st.success("All extracted invoices have complete Origin and Destination data.")

        with st.expander("📑 Duplicate Invoices Detected", expanded=False):
            st.write("Invoices sharing the same Vendor Name and Invoice Number:")
            if not duplicates_df.empty:
                duplicates_df['extracted_total_M'] = duplicates_df['extracted_total'].apply(lambda x: f"${x/1_000_000:,.2f}M")
                st.dataframe(duplicates_df[['vendor_name', 'invoice_number', 'file_name', 'extracted_total_M', 'extraction_date']], use_container_width=True, hide_index=True)
            else:
                st.success("No duplicate invoices detected in the master database.")
                
        with st.expander("📍 Vendor Logistics Map", expanded=False):
            st.write("Aggregated unique origins and destinations per vendor:")
            def get_unique_clean(series):
                vals = [str(x) for x in series if pd.notna(x) and str(x) not in missing_flags]
                return " | ".join(sorted(list(set(vals)))) if vals else "⚠️ Missing"

            df_routes = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])]
            vendor_routes = df_routes.groupby('vendor_name').agg({
                'id': 'count', 'origin': get_unique_clean, 'destination': get_unique_clean
            }).reset_index().rename(columns={
                'vendor_name': 'Vendor Name', 'id': 'Invoice Count', 'origin': 'Unique Origins', 'destination': 'Unique Destinations'
            }).sort_values(by='Invoice Count', ascending=False)
            
            st.dataframe(vendor_routes, use_container_width=True, hide_index=True)

# ---------------------------------------------------------
# TAB 5: SYSTEM & DRIFT ANALYSIS
# ---------------------------------------------------------
with tab_system:
    st.header("🤖 System & Model Drift Analysis")
    df_audit = fetch_audit_data()
    
    if df_audit.empty:
        st.info("📭 No data available. Process some invoices to track system drift.")
    else:
        if 'clean_supplier' in df_audit.columns:
            df_audit['vendor_name'] = df_audit['clean_supplier'].fillna(df_audit['vendor_name'].apply(standardize_vendor))
        else:
            df_audit['vendor_name'] = df_audit['vendor_name'].apply(standardize_vendor)
        
        df_audit['extraction_date_dt'] = pd.to_datetime(df_audit['extraction_date'])
        df_audit['datetime'] = pd.to_datetime(df_audit['extraction_date'] + ' ' + df_audit['extraction_time'])
        df_audit = df_audit.sort_values('datetime')
        df_audit['is_critical_error'] = df_audit['status'].str.contains("CRITICAL ERROR", na=False)
        df_audit['Clean_Status'] = df_audit['status'].apply(lambda x: 'PASS' if x == 'PASS' else 'FAIL')
        df_audit['processing_time_min'] = df_audit['processing_time'] / 60.0

        st.subheader("System Performance & Accuracy Metrics")
        sys1, sys2, sys3, sys4, sys5, sys6 = st.columns(6)
        
        total_invoices = len(df_audit)
        total_pages = df_audit['page_count'].sum() if 'page_count' in df_audit.columns else 0
        avg_proc_time_min = df_audit['processing_time_min'].mean() if 'processing_time_min' in df_audit.columns else 0.0
        sys_error_rate = (df_audit['is_critical_error'].sum() / total_invoices) * 100 if total_invoices > 0 else 0
        
        overall_pass = len(df_audit[df_audit['Clean_Status'] == 'PASS'])
        overall_fail = len(df_audit[df_audit['Clean_Status'] == 'FAIL'])
        overall_acc = (overall_pass / total_invoices) * 100 if total_invoices > 0 else 0
        fail_pct = (overall_fail / total_invoices) * 100 if total_invoices > 0 else 0
        
        math_pass = len(df_audit[df_audit['variance'] == 0.0])
        math_acc = (math_pass / total_invoices) * 100 if total_invoices > 0 else 0

        sys1.metric("Avg Proc Time", f"{avg_proc_time_min:.2f}m", "Per Doc")
        sys2.metric("API Failure Rate", f"{sys_error_rate:.1f}%", delta_color="inverse")
        sys3.metric("Overall Accuracy", f"{overall_acc:.1f}%")
        sys4.metric("Math Accuracy", f"{math_acc:.1f}%")
        sys5.metric("% Pass", f"{overall_acc:.1f}%")
        sys6.metric("% Fail", f"{fail_pct:.1f}%", delta_color="inverse")

        st.divider()

        s1, s2 = st.columns(2)
        with s1:
            st.subheader("Processing Speed Drift (Minutes)")
            fig_speed = px.line(df_audit, x='datetime', y='processing_time_min', markers=True, title='Document Processing Time', color_discrete_sequence=['#3498db'])
            fig_speed.update_layout(xaxis_title="Time", yaxis_title="Minutes")
            st.plotly_chart(fig_speed, use_container_width=True)

        with s2:
            st.subheader("Processing Volume by Day of Week")
            df_audit['day_of_week'] = df_audit['extraction_date_dt'].dt.day_name()
            day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            dow_counts = df_audit['day_of_week'].value_counts().reindex(day_order).reset_index()
            dow_counts.columns = ['Day', 'Volume']
            fig_dow = px.bar(dow_counts, x='Day', y='Volume', text_auto=True, color_discrete_sequence=['#1abc9c'])
            st.plotly_chart(fig_dow, use_container_width=True)
                
        st.divider()

        c3, c4 = st.columns(2)
        with c3:
            st.subheader("⚠️ Top Reasons for LLM Review")
            df_fails = df_audit[df_audit['Clean_Status'] == 'FAIL']
            if not df_fails.empty:
                reasons_series = df_fails['reason_for_review'].str.split(" | ").explode()
                
                normalized_reasons = []
                for r in reasons_series:
                    r_str = str(r)
                    if "Math Variance" in r_str: normalized_reasons.append("Amount Mismatch")
                    elif "Origin" in r_str: normalized_reasons.append("Origin Missing")
                    elif "Destination" in r_str: normalized_reasons.append("Destination Missing")
                    elif "UOM" in r_str: normalized_reasons.append("UOM Missing or Low Confidence")
                    elif "Invoice #" in r_str: normalized_reasons.append("Invoice Number Issue")
                    elif "0 Line Items" in r_str: normalized_reasons.append("No Line Items Found")
                    else: normalized_reasons.append("General Extraction Issue")
                
                reason_counts = pd.Series(normalized_reasons).value_counts().to_dict()
                phrase_counts = {k.replace(" ", "\xA0"): v for k, v in reason_counts.items()}
                
                if phrase_counts:
                    wordcloud = WordCloud(
                        width=800, 
                        height=400, 
                        background_color='white', 
                        colormap='viridis',
                        max_words=30
                    ).generate_from_frequencies(phrase_counts)
                    
                    fig, ax = plt.subplots(figsize=(8, 4))
                    ax.imshow(wordcloud, interpolation='bilinear')
                    ax.axis("off")
                    st.pyplot(fig)
                else:
                    st.info("No explicit reasons provided for the failures.")
            else:
                st.success("No failures to analyze.")

        with c4:
            st.subheader("💸 Absolute Variance Magnitude by Vendor")
            var_vendor = df_audit[~df_audit['vendor_name'].isin(['N/A', 'ERROR'])].groupby('vendor_name')['variance'].apply(lambda x: x.abs().sum()).reset_index()
            var_vendor = var_vendor[var_vendor['variance'] > 0].sort_values(by='variance', ascending=False).head(10)
            if not var_vendor.empty:
                fig_var_vendor = px.bar(var_vendor, x='variance', y='vendor_name', orientation='h', text_auto=True, color_discrete_sequence=['#e74c3c'])
                fig_var_vendor.update_layout(yaxis={'categoryorder':'total ascending'})
                st.plotly_chart(fig_var_vendor, use_container_width=True)
            else:
                st.success("No variance detected across any vendors.")
                
        st.divider()
            
        st.subheader("⚙️ Database Management")
        st.write("Use these controls to maintain system health and clear test data.")
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        with col_btn1:
            if st.button("🗑️ Remove Duplicates", use_container_width=True):
                dialog_remove_duplicates()
        with col_btn2:
            if st.button("⚠️ Wipe Database", type="secondary", use_container_width=True):
                dialog_wipe_db()
